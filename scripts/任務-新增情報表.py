"""
Add '任務情報' sheet to 萬法同歸_設計文件.xlsx
Contains all MainQuest data from Ch0-Ch8
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

XLSX_PATH = r"C:\Consilience\docs\萬法同歸_設計文件.xlsx"
SHEET_NAME = "任務情報"

# Column headers
HEADERS = [
    "Key",
    "Title",
    "RawTitle",
    "Difficulty",
    "From",
    "Location",
    "Description",
    "Objectives",
    "InitVisible",
    "Rewards",
    "Subtext",
    "Quote",
]

# Quest data for all 9 chapters (Ch0-Ch8)
QUESTS = [
    {
        "Key": "MainQuest_0_黃裳典籍",
        "Title": r"\i[3405]\c[14]黃裳遺址\c[0]的祕密",
        "RawTitle": "黃裳遺址的祕密",
        "Difficulty": "★",
        "From": "墨汐若",
        "Location": "黃裳遺址",
        "Description": (
            "杭州近郊深山中的古老遺址，\n"
            "相傳封存了上古武學總綱——\n"
            r"\c[2]黃裳典籍\c[0]。" + "\n"
            r"但\c[14]帝國\c[0]的人已搶先一步。" + "\n"
            "前往遺址深處，搶在他們之前找出真相。"
        ),
        "Objectives": (
            "1. 進入黃裳遺址\n"
            "2. 調查甬道中的帝國痕跡\n"
            "3. 探索中庭石室\n"
            "4. 深入遺址密道\n"
            "5. 調查地洞壁畫\n"
            "6. 擊退帝國斥候與魔物\n"
            "7. 分析鐘塔遺物與理式文件"
        ),
        "InitVisible": 1,
        "Rewards": "—",
        "Subtext": "被鑿去的符號之下，\n有些東西從未沉睡。",
        "Quote": "典籍已散，道猶在否？",
    },
    {
        "Key": "MainQuest_1_劃月風雲",
        "Title": r"\c[2]劃月風雲\c[0]",
        "RawTitle": "劃月風雲",
        "Difficulty": "★★",
        "From": "東方啟",
        "Location": "鎮江·聽風吟",
        "Description": (
            r"\c[18]((魑魅陰指))\c[0]殘勁侵蝕鎮江武者，" + "\n"
            r"殘勁中混有\c[2]((瑪那))\c[0]痕跡。" + "\n"
            "追查源頭，解救傷者。"
        ),
        "Objectives": (
            "1. 前往鎮江聽風吟\n"
            "2. 了解陰指殘勁傷者\n"
            "3. 發現殘勁中的瑪那\n"
            "4. 前往森林城鎮調查\n"
            "5. 擊退陰紋魔物\n"
            "6. 前往西域帝國海關站"
        ),
        "InitVisible": 1,
        "Rewards": "—",
        "Subtext": "陰指與瑪那的融合——\n誰在暗中操弄禁忌？",
        "Quote": "不格而毀，非道也。",
    },
    {
        "Key": "MainQuest_2_西域來風",
        "Title": r"\c[2]西域來風\c[0]",
        "RawTitle": "西域來風",
        "Difficulty": "★★★",
        "From": "東方啟",
        "Location": "西域·帝國海關站",
        "Description": (
            r"追查\c[18]((魑魅陰指))\c[0]的線索" + "\n"
            "指向西域邊境。\n"
            "帝國海關站中封存\n"
            "著疑似陰指的殘篇。"
        ),
        "Objectives": (
            "1. 通過帝國入境哨站\n"
            "2. 潛入海關站倉庫區\n"
            "3. 發現 Project Chimera 文件\n"
            "4. 救出被囚禁的黃沙族人\n"
            "5. 穿越黃沙古道\n"
            "6. 找到古道盡頭的線索"
        ),
        "InitVisible": 1,
        "Rewards": "—",
        "Subtext": (
            r"帝國的理性，\c[6]((求真))\c[0]；" + "\n"
            r"黃沙族的本能，\c[6]((求善))\c[0]。" + "\n"
            "兩者是否能共行？"
        ),
        "Quote": "土地不可測量。",
    },
    {
        "Key": "MainQuest_3_仙池寒劍",
        "Title": r"\c[2]仙池寒劍\c[0]",
        "RawTitle": "仙池寒劍",
        "Difficulty": "★★★",
        "From": "楊古晨",
        "Location": "仙池劍派",
        "Description": (
            r"循陰指線索抵達\c[14]仙池劍派\c[0]，" + "\n"
            r"探查\c[14]仙池深淵\c[0]的異象。" + "\n"
            "帝國的手，\n"
            "已伸入劍派深處。"
        ),
        "Objectives": (
            "1. 抵達江陵仙池山道\n"
            "2. 拜訪掌門龍玉\n"
            "3. 探查仙池深淵的異象\n"
            "4. 擊退仙池畸變體\n"
            "5. 擊退殷染幽\n"
            "6. 查明司徒長生的密謀\n"
            "7. 前往仙潭崖"
        ),
        "InitVisible": 1,
        "Rewards": "—",
        "Subtext": (
            r"「有用的」劍法留存了，\c[6]((求善))\c[0]的劍魂卻不知去向。" + "\n"
            "什麼才是真正的保全？"
        ),
        "Quote": "劍鋒折處，方見劍心。",
    },
    {
        "Key": "MainQuest_4_逍遙雲霧",
        "Title": r"\c[2]逍遙雲霧\c[0]",
        "RawTitle": "逍遙雲霧",
        "Difficulty": "★★★",
        "From": "談笑",
        "Location": "鯤鵬山·逍遙山門",
        "Description": (
            "仙池的封印不是孤例。\n"
            r"循談笑指引，前往鯤鵬山\c[14]逍遙山門\c[0]，" + "\n"
            r"向宗主\c[8]瑤琴劍\c[0]請教封印之事。" + "\n"
            "山中地脈正在衰退，\n"
            "帝國的手已伸入大地根基。"
        ),
        "Objectives": (
            "1. 前往鯤鵬山逍遙山門\n"
            "2. 與宗主瑤琴劍會面\n"
            "3. 前往觀測崖調查地脈異象\n"
            "4. 探索山中祕洞\n"
            "5. 揭開混沌封印的秘密\n"
            "6. 見證瑤琴劍的決定"
        ),
        "InitVisible": 1,
        "Rewards": "—",
        "Subtext": (
            "逍遙不是逃避。\n"
            r"當大地的根基被抽空，\c[6]((逍遙))\c[0]的精神也需要腳踩實地。"
        ),
        "Quote": "逍遙遊盡頭，仍需腳踩大地。",
    },
    {
        "Key": "MainQuest_5_梅莊庇護",
        "Title": r"\c[2]梅莊庇護\c[0]",
        "RawTitle": "梅莊庇護",
        "Difficulty": "★★★★",
        "From": "墨汐若",
        "Location": "梅莊",
        "Description": (
            "暫居梅莊休整。\n"
            r"帝國「\c[10]身分普查\c[0]」逼近，" + "\n"
            "無名丐帶回鐘塔的爆炸性情報。\n"
            "需要潛入總督府取得實驗記錄。"
        ),
        "Objectives": (
            "1. 抵達梅莊\n"
            "2. 聽取無名丐的情報\n"
            "3. 潛入帝國總督府\n"
            "4. 取得鐘塔實驗記錄\n"
            "5. 安全撤退\n"
            "6. 應對陰帥異象\n"
            "7. 撤離梅莊"
        ),
        "InitVisible": 1,
        "Rewards": "—",
        "Subtext": (
            r"「\c[6]兼愛\c[0]」在帝國眼中是非法的。" + "\n"
            "當收容本身成了罪，善良還有容身之處嗎？"
        ),
        "Quote": "梅花香自苦寒來。",
    },
    {
        "Key": "MainQuest_6_陰山暗影",
        "Title": r"\c[2]陰山暗影\c[0]",
        "RawTitle": "陰山暗影",
        "Difficulty": "★★★★★",
        "From": "湮菲花",
        "Location": "陰山草堂",
        "Description": (
            "梅莊淪陷後轉往陰山草堂。\n"
            "七霜揭示藥林焚毀的真相，\n"
            "藍靜冥帶來重傷的殷染幽，\n"
            "劉靜靜帶來鐘塔核心情報。\n"
            "戾神的出現將危機\n"
            "從政治層面推向存在層面。"
        ),
        "Objectives": (
            "1. 抵達陰山草堂\n"
            "2. 安頓傷者\n"
            "3. 調查長江枯林\n"
            "4. 治療殷染幽\n"
            "5. 應對戾神異象\n"
            "6. 接收劉靜靜的情報\n"
            "7. 探索地下遺跡\n"
            "8. 取得靈寶經殘卷與甘珠爾\n"
            "9. 決定突入鐘塔"
        ),
        "InitVisible": 1,
        "Rewards": "—",
        "Subtext": (
            "全篇最黑暗的轉折——\n"
            "藥林焚毀、五仙教滅門、戾神現世。\n"
            "但黑暗中也有轉機：\n"
            r"\c[6]入魔並非不可逆，焚毀之後能有新生。\c[0]"
        ),
        "Quote": "知其白，守其黑。",
    },
    {
        "Key": "MainQuest_7_鐘塔審判",
        "Title": r"\c[2]鐘塔審判\c[0]",
        "RawTitle": "鐘塔審判",
        "Difficulty": "★★★★★",
        "From": "東方啟",
        "Location": "鐘塔研究院",
        "Description": (
            "正面突入帝國鐘塔研究院。\n"
            "黃凱竹提供密道，\n"
            "龍玉與司徒長生意外合流。\n"
            "深入鐘塔的過程中，\n"
            "將親眼目睹帝國的一切。\n"
            "在最深處，做出最終抉擇。"
        ),
        "Objectives": (
            "1. 與黃凱竹會合\n"
            "2. 進入密道\n"
            "3. 穿越研究區\n"
            "4. 面對奇美拉軍團\n"
            "5. 抵達鐘塔核心\n"
            "6. 面對不可名狀之物的影子\n"
            "7. 做出最終抉擇"
        ),
        "InitVisible": 1,
        "Rewards": "—",
        "Subtext": (
            "鐘塔是全篇衝突的物質化——\n"
            "將「道」拆解為公式，\n"
            "是帝國「線性理性」的終極展現。\n"
            r"\c[6]面對那片虛空，所有法門都同樣蒼白。\c[0]"
        ),
        "Quote": "雖千萬人，吾往矣。",
    },
    {
        "Key": "MainQuest_8_萬法同歸",
        "Title": r"\c[2]萬法同歸\c[0]",
        "RawTitle": "萬法同歸",
        "Difficulty": "★★★★★",
        "From": "—",
        "Location": "依選擇分歧",
        "Description": (
            "鐘塔崩塌，混沌之源甦醒。\n"
            "依據第七章的選擇，\n"
            "走向三條截然不同的道路。\n"
            "每一條路都有代價，\n"
            "每一條路都有希望。"
        ),
        "Objectives": (
            "1. 撤離崩塌的鐘塔\n"
            "2. 依選擇執行最終行動\n"
            "3. 面對結局"
        ),
        "InitVisible": 1,
        "Rewards": "—",
        "Subtext": (
            r"\c[6]萬法不歸神佛，終歸自覺。\c[0]" + "\n"
            "在所有法門都失效的邊界上，\n"
            "唯一能依靠的，\n"
            "是自己那顆不願被消解的心。"
        ),
        "Quote": "何期自性，能生萬法。",
    },
]

def main():
    wb = openpyxl.load_workbook(XLSX_PATH)

    # Remove existing sheet if present
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]

    ws = wb.create_sheet(SHEET_NAME)

    # Styles
    header_font = Font(name="Microsoft JhengHei", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2B3A4A", end_color="2B3A4A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="Microsoft JhengHei", size=10)
    data_align = Alignment(vertical="top", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="AAAAAA"),
        right=Side(style="thin", color="AAAAAA"),
        top=Side(style="thin", color="AAAAAA"),
        bottom=Side(style="thin", color="AAAAAA"),
    )

    alt_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")

    # Write headers
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data
    for row_idx, quest in enumerate(QUESTS, 2):
        for col_idx, key in enumerate(HEADERS, 1):
            value = quest.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Column widths
    col_widths = {
        "A": 28,   # Key
        "B": 30,   # Title
        "C": 18,   # RawTitle
        "D": 10,   # Difficulty
        "E": 10,   # From
        "F": 22,   # Location
        "G": 45,   # Description
        "H": 40,   # Objectives
        "I": 12,   # InitVisible
        "J": 12,   # Rewards
        "K": 45,   # Subtext
        "L": 24,   # Quote
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:L{len(QUESTS) + 1}"

    wb.save(XLSX_PATH)
    print(f"Done! Sheet '{SHEET_NAME}' added with {len(QUESTS)} quests.")
    print(f"Saved to: {XLSX_PATH}")


if __name__ == "__main__":
    main()
