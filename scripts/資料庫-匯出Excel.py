# -*- coding: utf-8 -*-
"""
將 RPG Maker MZ 遊戲資料完整匯出至 萬法同歸_資料庫.xlsx
匯出範圍：Actors, Classes, Skills, Items, Weapons, Armors, Enemies, States,
          MapInfos, Troops, Elements, WeaponTypes, ArmorTypes, EquipTypes,
          SkillTypes, Params, BasicTerms, Switches, Variables
"""
import json
import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

os.environ['PYTHONUTF8'] = '1'

DATA_DIR = r'C:\Consilience\Consilience\data'
OUT_PATH = r'C:\Consilience\docs\萬法同歸_資料庫.xlsx'


def load_json(filename):
    with open(os.path.join(DATA_DIR, filename), 'r', encoding='utf-8') as f:
        return json.load(f)


def strip_codes(text):
    """移除 RPG Maker 色碼 \\c[N] 與圖標碼 \\I[N]"""
    if not text:
        return ''
    text = re.sub(r'\\[cCiI]\[\d+\]', '', text)
    return text.strip()


# ── 樣式 ─────────────────────────────────────────────
HEADER_FONT = Font(name='Microsoft JhengHei', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_FONT = Font(name='Microsoft JhengHei', size=10)
CELL_ALIGN = Alignment(vertical='top', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)


def style_sheet(ws, headers, col_widths=None):
    """套用標題樣式並設定欄寬"""
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    if col_widths:
        for col_idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions


def write_row(ws, row_idx, values):
    for col_idx, v in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=v)
        cell.font = CELL_FONT
        cell.alignment = CELL_ALIGN
        cell.border = THIN_BORDER


# ── scope 對照表 ──────────────────────────────────────
SCOPE_MAP = {
    0: '無', 1: '敵單', 2: '敵全', 3: '敵1隨機', 4: '敵2隨機',
    5: '敵3隨機', 6: '敵4隨機', 7: '我單', 8: '我全',
    9: '我方死亡單', 10: '我方死亡全', 11: '使用者', 12: '敵單(即死)',
    13: '我全(含死)', 14: '敵全(含死)',
}
DMG_TYPE_MAP = {
    0: '無', 1: 'HP傷害', 2: 'MP傷害', 3: 'HP回復',
    4: 'MP回復', 5: 'HP吸收', 6: 'MP吸收',
}
OCCASION_MAP = {0: '隨時', 1: '僅戰鬥', 2: '僅選單', 3: '不可用'}


def build_actors(wb):
    actors = load_json('Actors.json')
    ws = wb.create_sheet('Actors')
    headers = ['ID', '名稱', '稱號', '職業ID', '初始等級', '最大等級',
               '行走圖', '臉圖', '個人簡介', '備註(Note)']
    style_sheet(ws, headers, [5, 10, 10, 8, 8, 8, 18, 18, 40, 40])
    row = 2
    for a in actors:
        if a is None or not a.get('name', '').strip():
            continue
        write_row(ws, row, [
            a['id'], a['name'], a.get('nickname', ''),
            a.get('classId', ''), a.get('initialLevel', ''),
            a.get('maxLevel', ''), a.get('characterName', ''),
            a.get('faceName', ''), a.get('profile', ''),
            a.get('note', ''),
        ])
        row += 1


def build_classes(wb):
    classes = load_json('Classes.json')
    ws = wb.create_sheet('Classes')
    headers = ['ID', '名稱', '裝備類型', '技能習得', '備註(Note)']
    style_sheet(ws, headers, [5, 18, 20, 40, 40])
    row = 2
    for c in classes:
        if c is None or not c.get('name', '').strip():
            continue
        # 裝備類型
        etypes = ', '.join(str(e) for e in c.get('traits', [])
                           if isinstance(e, dict) and e.get('code') == 51)
        # 技能習得列表
        learnings = c.get('learnings', [])
        learn_str = '; '.join(
            f"Lv{l['level']}→Skill#{l['skillId']}" for l in learnings
        ) if learnings else ''
        write_row(ws, row, [
            c['id'], c['name'], etypes, learn_str, c.get('note', ''),
        ])
        row += 1


def build_skills(wb):
    skills = load_json('Skills.json')
    ws = wb.create_sheet('Skills')
    headers = ['ID', '名稱', '說明', '技能類型ID', '範圍', 'MP消耗', 'TP消耗',
               'TP獲得', '傷害類型', '傷害公式', '屬性ID', '變異',
               '命中類型', '連續次數', '速度補正', '成功率', '效果', '備註(Note)']
    style_sheet(ws, headers, [5, 14, 30, 10, 8, 8, 8, 8, 8, 25, 8, 6, 8, 8, 8, 8, 30, 30])
    row = 2
    for s in skills:
        if s is None or not s.get('name', '').strip():
            continue
        dmg = s.get('damage', {})
        effects = s.get('effects', [])
        eff_str = '; '.join(
            f"code={e['code']} id={e.get('dataId','')} v1={e.get('value1','')} v2={e.get('value2','')}"
            for e in effects
        ) if effects else ''
        write_row(ws, row, [
            s['id'], s['name'], strip_codes(s.get('description', '')),
            s.get('stypeId', ''),
            SCOPE_MAP.get(s.get('scope', 0), str(s.get('scope', 0))),
            s.get('mpCost', 0), s.get('tpCost', 0), s.get('tpGain', 0),
            DMG_TYPE_MAP.get(dmg.get('type', 0), str(dmg.get('type', 0))),
            dmg.get('formula', ''), dmg.get('elementId', ''),
            dmg.get('variance', ''), s.get('hitType', ''),
            s.get('repeats', 1), s.get('speed', 0),
            s.get('successRate', 100), eff_str, s.get('note', ''),
        ])
        row += 1


def build_items(wb):
    items = load_json('Items.json')
    ws = wb.create_sheet('Items')
    headers = ['ID', '名稱', '說明', '道具類型', '價格', '消耗品',
               '範圍', '速度補正', '成功率', '效果', '備註(Note)']
    style_sheet(ws, headers, [5, 14, 35, 10, 8, 8, 10, 8, 8, 35, 30])
    itype_map = {1: '一般', 2: '重要', 3: '隱藏A', 4: '隱藏B'}
    row = 2
    for it in items:
        if it is None or not it.get('name', '').strip():
            continue
        effects = it.get('effects', [])
        eff_str = '; '.join(
            f"code={e['code']} id={e.get('dataId','')} v1={e.get('value1','')} v2={e.get('value2','')}"
            for e in effects
        ) if effects else ''
        write_row(ws, row, [
            it['id'], it['name'], strip_codes(it.get('description', '')),
            itype_map.get(it.get('itypeId', 1), str(it.get('itypeId', 1))),
            it.get('price', 0),
            '否' if it.get('consumable', True) is False else '是',
            SCOPE_MAP.get(it.get('scope', 0), str(it.get('scope', 0))),
            it.get('speed', 0), it.get('successRate', 100),
            eff_str, it.get('note', ''),
        ])
        row += 1


def build_weapons(wb):
    weapons = load_json('Weapons.json')
    system = load_json('System.json')
    wtype_names = system.get('weaponTypes', [])
    ws = wb.create_sheet('Weapons')
    headers = ['ID', '名稱', '說明', '武器類型', '價格',
               'MHP', 'MMP', '外功', '外防', '內功', '內防', '輕功', '運勢',
               '特性', '備註(Note)']
    style_sheet(ws, headers, [5, 14, 30, 10, 8, 8, 8, 8, 8, 8, 8, 8, 8, 30, 30])
    row = 2
    for w in weapons:
        if w is None or not w.get('name', '').strip():
            continue
        p = w.get('params', [0]*8)
        wt = w.get('wtypeId', 0)
        wt_name = strip_codes(wtype_names[wt]) if wt < len(wtype_names) else str(wt)
        traits = w.get('traits', [])
        trait_str = '; '.join(
            f"code={t['code']} id={t.get('dataId','')} val={t.get('value','')}"
            for t in traits
        ) if traits else ''
        write_row(ws, row, [
            w['id'], w['name'], strip_codes(w.get('description', '')),
            wt_name, w.get('price', 0),
            p[0], p[1], p[2], p[3], p[4], p[5], p[6], p[7],
            trait_str, w.get('note', ''),
        ])
        row += 1


def build_armors(wb):
    armors = load_json('Armors.json')
    system = load_json('System.json')
    atype_names = system.get('armorTypes', [])
    etype_names = system.get('equipTypes', [])
    ws = wb.create_sheet('Armors')
    headers = ['ID', '名稱', '說明', '裝備類型', '護甲類型', '價格',
               'MHP', 'MMP', '外功', '外防', '內功', '內防', '輕功', '運勢',
               '特性', '備註(Note)']
    style_sheet(ws, headers, [5, 14, 30, 10, 10, 8, 8, 8, 8, 8, 8, 8, 8, 8, 30, 30])
    row = 2
    for a in armors:
        if a is None or not a.get('name', '').strip():
            continue
        p = a.get('params', [0]*8)
        et = a.get('etypeId', 0)
        at = a.get('atypeId', 0)
        et_name = strip_codes(etype_names[et]) if et < len(etype_names) else str(et)
        at_name = strip_codes(atype_names[at]) if at < len(atype_names) else str(at)
        traits = a.get('traits', [])
        trait_str = '; '.join(
            f"code={t['code']} id={t.get('dataId','')} val={t.get('value','')}"
            for t in traits
        ) if traits else ''
        write_row(ws, row, [
            a['id'], a['name'], strip_codes(a.get('description', '')),
            et_name, at_name, a.get('price', 0),
            p[0], p[1], p[2], p[3], p[4], p[5], p[6], p[7],
            trait_str, a.get('note', ''),
        ])
        row += 1


def build_enemies(wb):
    enemies = load_json('Enemies.json')
    ws = wb.create_sheet('Enemies')
    headers = ['ID', '名稱', '戰鬥圖', '經驗值', '金幣',
               'MHP', 'MMP', '外功', '外防', '內功', '內防', '輕功', '運勢',
               '掉落物品', '行動模式', '特性', '備註(Note)']
    style_sheet(ws, headers, [5, 14, 18, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 25, 30, 30, 30])
    drop_kind = {0: '無', 1: '道具', 2: '武器', 3: '防具'}
    row = 2
    for e in enemies:
        if e is None or not e.get('name', '').strip():
            continue
        p = e.get('params', [0]*8)
        drops = e.get('dropItems', [])
        drop_str = '; '.join(
            f"{drop_kind.get(d['kind'],'?')}#{d['dataId']}(1/{d['denominator']})"
            for d in drops if d.get('kind', 0) != 0
        ) if drops else ''
        actions = e.get('actions', [])
        act_str = '; '.join(
            f"Skill#{a['skillId']}(rating={a['rating']})"
            for a in actions
        ) if actions else ''
        traits = e.get('traits', [])
        trait_str = '; '.join(
            f"code={t['code']} id={t.get('dataId','')} val={t.get('value','')}"
            for t in traits
        ) if traits else ''
        write_row(ws, row, [
            e['id'], e['name'], e.get('battlerName', ''),
            e.get('exp', 0), e.get('gold', 0),
            p[0], p[1], p[2], p[3], p[4], p[5], p[6], p[7],
            drop_str, act_str, trait_str, e.get('note', ''),
        ])
        row += 1


def build_states(wb):
    states = load_json('States.json')
    ws = wb.create_sheet('States')
    headers = ['ID', '名稱', '圖標', '優先度', '限制',
               '自動解除', '最小回合', '最大回合',
               '步數解除率', '戰鬥結束解除', '傷害解除',
               '特性', '備註(Note)']
    style_sheet(ws, headers, [5, 14, 6, 8, 8, 10, 8, 8, 10, 12, 10, 30, 30])
    restriction_map = {0: '無', 1: '攻擊敵人', 2: '攻擊任意', 3: '攻擊友軍', 4: '不行動'}
    removal_map = {0: '無', 1: '行動結束', 2: '回合結束'}
    row = 2
    for st in states:
        if st is None or not st.get('name', '').strip():
            continue
        traits = st.get('traits', [])
        trait_str = '; '.join(
            f"code={t['code']} id={t.get('dataId','')} val={t.get('value','')}"
            for t in traits
        ) if traits else ''
        write_row(ws, row, [
            st['id'], st['name'], st.get('iconIndex', 0),
            st.get('priority', 0),
            restriction_map.get(st.get('restriction', 0), ''),
            removal_map.get(st.get('autoRemovalTiming', 0), ''),
            st.get('minTurns', 0), st.get('maxTurns', 0),
            st.get('stepsToRemove', 0),
            '是' if st.get('removeAtBattleEnd', False) else '否',
            '是' if st.get('removeByDamage', False) else '否',
            trait_str, st.get('note', ''),
        ])
        row += 1


def build_map_infos(wb):
    maps = load_json('MapInfos.json')
    ws = wb.create_sheet('MapInfos')
    headers = ['ID', '名稱', '母地圖ID', '排序']
    style_sheet(ws, headers, [5, 20, 10, 6])
    row = 2
    for m in maps:
        if m is None or not m.get('name', '').strip():
            continue
        write_row(ws, row, [
            m['id'], m['name'], m.get('parentId', 0), m.get('order', 0),
        ])
        row += 1


def build_troops(wb):
    troops = load_json('Troops.json')
    ws = wb.create_sheet('Troops')
    headers = ['ID', '名稱', '成員(敵人ID)']
    style_sheet(ws, headers, [5, 25, 30])
    row = 2
    for t in troops:
        if t is None or not t.get('name', '').strip():
            continue
        members = ', '.join(str(m['enemyId']) for m in t.get('members', []))
        write_row(ws, row, [t['id'], t['name'], members])
        row += 1


def build_system_lists(wb):
    system = load_json('System.json')

    # Elements
    ws = wb.create_sheet('Elements')
    headers = ['ID', '屬性名稱']
    style_sheet(ws, headers, [5, 20])
    for i, name in enumerate(system.get('elements', [])):
        if i == 0:
            continue
        write_row(ws, i + 1, [i, strip_codes(name)])

    # WeaponTypes
    ws = wb.create_sheet('WeaponTypes')
    headers = ['ID', '武器類型']
    style_sheet(ws, headers, [5, 15])
    for i, name in enumerate(system.get('weaponTypes', [])):
        if i == 0:
            continue
        write_row(ws, i + 1, [i, strip_codes(name)])

    # ArmorTypes
    ws = wb.create_sheet('ArmorTypes')
    headers = ['ID', '護甲類型']
    style_sheet(ws, headers, [5, 15])
    for i, name in enumerate(system.get('armorTypes', [])):
        if i == 0:
            continue
        write_row(ws, i + 1, [i, strip_codes(name)])

    # EquipTypes
    ws = wb.create_sheet('EquipTypes')
    headers = ['ID', '裝備類型']
    style_sheet(ws, headers, [5, 15])
    for i, name in enumerate(system.get('equipTypes', [])):
        if i == 0:
            continue
        write_row(ws, i + 1, [i, strip_codes(name)])

    # SkillTypes
    ws = wb.create_sheet('SkillTypes')
    headers = ['ID', '技能類型']
    style_sheet(ws, headers, [5, 15])
    for i, name in enumerate(system.get('skillTypes', [])):
        if i == 0:
            continue
        write_row(ws, i + 1, [i, strip_codes(name)])

    # Params
    ws = wb.create_sheet('Params')
    headers = ['ID', '參數名稱']
    style_sheet(ws, headers, [5, 15])
    params = system.get('terms', {}).get('params', [])
    for i, name in enumerate(params):
        write_row(ws, i + 2, [i, strip_codes(name)])

    # BasicTerms
    ws = wb.create_sheet('BasicTerms')
    headers = ['ID', '基本用語']
    style_sheet(ws, headers, [5, 15])
    basic = system.get('terms', {}).get('basic', [])
    for i, name in enumerate(basic):
        write_row(ws, i + 2, [i, strip_codes(name)])

    # Switches (only named ones)
    ws = wb.create_sheet('Switches')
    headers = ['ID', '開關名稱']
    style_sheet(ws, headers, [5, 25])
    row = 2
    for i, name in enumerate(system.get('switches', [])):
        if name and name.strip():
            write_row(ws, row, [i, name])
            row += 1

    # Variables (only named ones)
    ws = wb.create_sheet('Variables')
    headers = ['ID', '變數名稱']
    style_sheet(ws, headers, [5, 25])
    row = 2
    for i, name in enumerate(system.get('variables', [])):
        if name and name.strip():
            write_row(ws, row, [i, name])
            row += 1


def main():
    wb = openpyxl.Workbook()
    # 刪除預設 Sheet
    wb.remove(wb.active)

    print('匯出 Actors...')
    build_actors(wb)
    print('匯出 Classes...')
    build_classes(wb)
    print('匯出 Skills...')
    build_skills(wb)
    print('匯出 Items...')
    build_items(wb)
    print('匯出 Weapons...')
    build_weapons(wb)
    print('匯出 Armors...')
    build_armors(wb)
    print('匯出 Enemies...')
    build_enemies(wb)
    print('匯出 States...')
    build_states(wb)
    print('匯出 MapInfos...')
    build_map_infos(wb)
    print('匯出 Troops...')
    build_troops(wb)
    print('匯出 System Lists...')
    build_system_lists(wb)

    wb.save(OUT_PATH)
    print(f'\n完成！已匯出至 {OUT_PATH}')
    # 統計
    for ws in wb.worksheets:
        print(f'  {ws.title}: {ws.max_row - 1} 筆資料')


if __name__ == '__main__':
    main()
