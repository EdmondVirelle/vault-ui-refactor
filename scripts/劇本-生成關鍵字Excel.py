"""Generate keywords Excel from 序章_keywords.md"""
import json
import re
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Color code mapping for SceneGlossary
COLOR_MAP = {
    0: "白色 (預設)",
    2: "橙色 (道具/典籍)",
    6: "黃色 (概念/哲學)",
    8: "灰色 (稱號)",
    10: "紅色 (危險/封印)",
    14: "青色 (地點/組織)",
    18: "紫色 (禁忌/混沌)",
    20: "綠色 (戰鬥/系統)",
}

def parse_color_code(text_str: str) -> str:
    """Extract color code number from \\c[N]"""
    m = re.search(r'\\c\[(\d+)\]', text_str)
    return int(m.group(1)) if m else 0

def get_fill_color(code: int) -> str:
    """Return hex color for Excel cell fill based on RPG Maker color code."""
    mapping = {
        2:  "FFA500",  # orange
        6:  "FFD700",  # gold/yellow
        8:  "999999",  # gray
        10: "FF4444",  # red
        14: "00CED1",  # dark turquoise
        18: "9966CC",  # purple
        20: "66CC66",  # green
    }
    return mapping.get(code, None)

# Read keywords from _array.txt (properly escaped JSON array)
keywords_path = Path(__file__).parent.parent / "consilience-writer" / "序章_keywords_array.txt"
keywords = []

raw = keywords_path.read_text(encoding="utf-8").strip()
# Parse the outer JSON array — each element is a JSON string containing an object
entries = json.loads(raw)

for entry_str in entries:
    # entry_str contains \c[N] which isn't valid JSON when re-parsed,
    # so extract fields with regex instead
    kw_m = re.search(r'"Keyword:str"\s*:\s*"([^"]*)"', entry_str)
    text_m = re.search(r'"Text:str"\s*:\s*"([^"]*)"', entry_str)
    tip_m = re.search(r'"Tooltip:json"\s*:\s*"(.*)"', entry_str)
    if not (kw_m and text_m and tip_m):
        continue

    kw = kw_m.group(1)
    text = text_m.group(1)
    # Tooltip is double-escaped: strip outer quotes and unescape
    tip_raw = tip_m.group(1)
    # Remove wrapping escaped quotes: \"...\"
    tip_raw = re.sub(r'^\\?"', '', tip_raw)
    tip_raw = re.sub(r'\\?"$', '', tip_raw)
    tooltip = tip_raw.replace("\\n", "\n")

    color_code = parse_color_code(text)
    keywords.append((kw, text, tooltip, color_code))

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "序章"

# Header style
header_font = Font(name="Microsoft JhengHei", bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Headers
headers = ["Keyword", "Replacement Text", "Tooltip Text", "色碼", "色碼說明"]
col_widths = [18, 30, 55, 8, 22]

for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws.column_dimensions[chr(64 + col_idx)].width = width

# Data rows
data_font = Font(name="Microsoft JhengHei", size=11)
wrap_align = Alignment(vertical="top", wrap_text=True)

for row_idx, (kw, text, tooltip, color_code) in enumerate(keywords, 2):
    # Keyword
    c1 = ws.cell(row=row_idx, column=1, value=kw)
    c1.font = Font(name="Microsoft JhengHei", size=11, bold=True)
    c1.alignment = Alignment(vertical="top")
    c1.border = thin_border

    # Replacement Text
    c2 = ws.cell(row=row_idx, column=2, value=text)
    c2.font = data_font
    c2.alignment = Alignment(vertical="top")
    c2.border = thin_border

    # Tooltip (replace \n with actual newlines for readability)
    c3 = ws.cell(row=row_idx, column=3, value=tooltip)
    c3.font = data_font
    c3.alignment = wrap_align
    c3.border = thin_border

    # Color code number
    c4 = ws.cell(row=row_idx, column=4, value=color_code)
    c4.font = data_font
    c4.alignment = Alignment(horizontal="center", vertical="top")
    c4.border = thin_border
    fill_hex = get_fill_color(color_code)
    if fill_hex:
        c4.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
        # Use white text on dark backgrounds
        if color_code in (10, 18, 8):
            c4.font = Font(name="Microsoft JhengHei", size=11, color="FFFFFF")

    # Color description
    c5 = ws.cell(row=row_idx, column=5, value=COLOR_MAP.get(color_code, f"\\c[{color_code}]"))
    c5.font = data_font
    c5.alignment = Alignment(vertical="top")
    c5.border = thin_border

# Freeze top row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:E{len(keywords) + 1}"

output_path = Path(__file__).parent.parent / "docs" / "萬法同歸_關鍵字總表.xlsx"
wb.save(str(output_path))
print(f"Created: {output_path}")
print(f"Total keywords in 序章: {len(keywords)}")
