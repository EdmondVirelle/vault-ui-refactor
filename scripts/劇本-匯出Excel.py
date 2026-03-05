#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
把 consilience-writer 的「序章~第八章前綴檔」匯入「萬法同歸_設計文件.xlsx」。

核心規則：
1) 每章自動掃描 {章名}_*.md，分類自動判斷（劇本/任務/補充）
2) 主線與支線任務頁不再強制獨立，整併到各章「整合」頁
3) 任務編號可直接寫在劇本序號旁（例如 S1-033 + MainQuest_xxx）
   腳本會自動產生「任務索引（依劇本序號）」供檢查
4) 路人對話觸發矩陣獨立彙整到單一總表（序章到第八章）

相容舊參數：
  python 劇本-匯出Excel.py
  python 劇本-匯出Excel.py 1 2
  python 劇本-匯出Excel.py 0-3
  python 劇本-匯出Excel.py 1 --with-expansion
  python 劇本-匯出Excel.py --list
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(r"C:\Consilience")
WRITER_DIR = BASE_DIR / "consilience-writer"
WORKBOOK_PATH = BASE_DIR / "docs" / "萬法同歸_設計文件.xlsx"

CHAPTERS: list[tuple[int, str]] = [
    (0, "序章"),
    (1, "第一章"),
    (2, "第二章"),
    (3, "第三章"),
    (4, "第四章"),
    (5, "第五章"),
    (6, "第六章"),
    (7, "第七章"),
    (8, "第八章"),
]

EXPANSIONS: list[tuple[str, str, set[int]]] = [
    ("擴充_序章第一章", "擴充_序章與第一章.md", {0, 1}),
    ("擴充_第二章第三章", "擴充_第二章與第三章.md", {2, 3}),
    ("擴充_第四章第五章", "擴充_第四章與第五章.md", {4, 5}),
    ("擴充_第六章第七章", "擴充_第六章與第七章.md", {6, 7}),
    ("擴充_第八章", "擴充_第八章.md", {8}),
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Microsoft JhengHei", bold=True, color="FFFFFF", size=10)
CELL_FONT = Font(name="Microsoft JhengHei", size=10)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

MAIN_QUEST_SUFFIX = "_主線任務.md"
SIDE_QUEST_SUFFIX = "_支線任務.md"
NPC_MATRIX_SUFFIX = "_路人對話觸發矩陣.md"
NPC_AGGREGATE_SHEET = "序章到第八章_路人觸發對話"

SEQ_RE = re.compile(r"\bS\d+(?:-[A-Za-z0-9]+)+\b")
QUEST_RE = re.compile(r"\b(?:MainQuest|SubQuest)_[A-Za-z0-9_\-]+\b|\b(?:MQ|SQ)\d+(?:-\d+){2}\b")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="匯入序章~第八章前綴 Markdown 到設計文件 Excel（自動分類）")
    parser.add_argument("chapters", nargs="*", default=[], help="章節索引：0=序章，1..8=第一章..第八章，可用 0-3")
    parser.add_argument("--with-expansion", "-e", action="store_true", help="匯入擴充章節檔")
    parser.add_argument("--list", "-l", action="store_true", help="列出可用章節")
    parser.add_argument("--dry-run", action="store_true", help="只顯示動作，不寫入 Excel")
    return parser


def parse_chapter_args(args: list[str]) -> set[int] | None:
    if not args:
        return None
    out: set[int] = set()
    for raw in args:
        raw = raw.strip()
        if not raw:
            continue
        if "-" in raw:
            s, e = raw.split("-", 1)
            start = int(s)
            end = int(e)
            out.update(range(start, end + 1))
        else:
            out.add(int(raw))
    for idx in out:
        if idx < 0 or idx > 8:
            raise ValueError(f"章節索引超出範圍: {idx}")
    return out


def parse_markdown_tables(text: str) -> list[tuple[str, list[list[str]]]]:
    sections: list[tuple[str, list[list[str]]]] = []
    current_header = ""
    lines = text.splitlines()
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if line.startswith("## ") or line.startswith("### "):
            current_header = line.lstrip("#").strip()
            i += 1
            continue
        if line.startswith("|"):
            table_lines = [line]
            i += 1
            while i < len(lines) and lines[i].strip().startswith("|"):
                table_lines.append(lines[i].strip())
                i += 1
            table = parse_single_table(table_lines)
            if table:
                sections.append((current_header, table))
            continue
        i += 1
    return sections


def parse_single_table(lines: list[str]) -> list[list[str]]:
    if len(lines) < 2:
        return []
    rows: list[list[str]] = []
    placeholder = "\x00PIPE\x00"
    for idx, line in enumerate(lines):
        if idx == 1 and re.match(r"^\|[\s:\-|]+\|$", line):
            continue
        safe = line.replace(r"\|", placeholder)
        cells = [x.strip().replace(placeholder, r"\|") for x in safe.split("|")]
        if cells and cells[0] == "":
            cells = cells[1:]
        if cells and cells[-1] == "":
            cells = cells[:-1]
        if cells:
            rows.append(cells)
    return rows


def read_md_sections(path: Path) -> list[tuple[str, list[list[str]]]]:
    if not path.exists():
        return [("MISSING", [[f"找不到檔案: {path.name}"]])]
    text = path.read_text(encoding="utf-8")
    sections = parse_markdown_tables(text)
    if not sections:
        return [("EMPTY", [[f"檔案沒有 Markdown 表格: {path.name}"]])]
    return sections


def flatten_sections(sections: list[tuple[str, list[list[str]]]]) -> list[list[str]]:
    rows: list[list[str]] = []
    for header, table in sections:
        title = header if header else "未命名段落"
        rows.append([f"## {title}"])
        rows.extend(table)
        rows.append([])
    while rows and not any(str(c).strip() for c in rows[-1]):
        rows.pop()
    return normalize_rows(rows)


def normalize_rows(rows: list[list[str]]) -> list[list[str]]:
    if not rows:
        return rows
    width = max(len(r) for r in rows)
    return [r + [""] * (width - len(r)) for r in rows]


def style_sheet(ws, rows: list[list[str]]) -> None:
    if not rows:
        return
    for r_idx, row in enumerate(rows, start=1):
        is_header_row = r_idx == 1 or (row and str(row[0]).startswith("## "))
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = THIN_BORDER
            if is_header_row:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGN
            else:
                cell.font = CELL_FONT
                cell.alignment = CELL_ALIGN

    widths = [0] * len(rows[0])
    for row in rows:
        for i, v in enumerate(row):
            widths[i] = min(max(widths[i], len(str(v))), 70)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(10, min(70, w + 2))
    ws.freeze_panes = "A2"


def sheet_title(chapter_name: str, component: str) -> str:
    # Excel sheet <= 31 chars
    title = f"{chapter_name}_{component}"
    return title[:31]


def chapter_markdown_files(chapter_name: str) -> list[Path]:
    files = sorted(WRITER_DIR.glob(f"{chapter_name}_*.md"), key=lambda p: p.name)
    return [p for p in files if p.is_file()]


def classify_chapter_file(path: Path) -> str:
    name = path.name
    if name.endswith(NPC_MATRIX_SUFFIX):
        return "npc"
    if name.endswith(MAIN_QUEST_SUFFIX):
        return "main_quest"
    if name.endswith(SIDE_QUEST_SUFFIX):
        return "side_quest"
    if "任務" in name:
        return "quest_other"
    return "story_or_other"


def file_category_priority(category: str) -> int:
    priorities = {
        "story_or_other": 0,
        "main_quest": 1,
        "side_quest": 2,
        "quest_other": 3,
        "npc": 9,
    }
    return priorities.get(category, 8)


def section_title_for_file(path: Path, category: str) -> str:
    if category == "main_quest":
        prefix = "主線任務"
    elif category == "side_quest":
        prefix = "支線任務"
    elif category == "quest_other":
        prefix = "任務補充"
    else:
        prefix = "劇本/補充"
    return f"{prefix} | {path.name}"


def normalize_seq_for_sort(seq: str) -> tuple[int, list[object]]:
    parts: list[object] = []
    for token in re.split(r"(\d+)", seq):
        if not token:
            continue
        if token.isdigit():
            parts.append(int(token))
        else:
            parts.append(token)
    chap_match = re.search(r"S(\d+)", seq)
    chapter_num = int(chap_match.group(1)) if chap_match else 999
    return chapter_num, parts


def extract_sequence_quest_pairs(rows: list[list[str]], source_name: str) -> list[tuple[str, str, str]]:
    out: list[tuple[str, str, str]] = []
    seen: set[tuple[str, str, str]] = set()
    for row in rows:
        text = " | ".join(str(c) for c in row if str(c).strip())
        if not text:
            continue
        seqs = sorted(set(SEQ_RE.findall(text)))
        quests = sorted(set(QUEST_RE.findall(text)))
        if not seqs or not quests:
            continue
        for seq in seqs:
            for quest in quests:
                key = (seq, quest, source_name)
                if key in seen:
                    continue
                seen.add(key)
                out.append(key)
    return out


def build_chapter_rows_and_quests(chapter_name: str) -> tuple[list[list[str]], list[tuple[str, str, str]]]:
    files = chapter_markdown_files(chapter_name)
    files = sorted(files, key=lambda p: (file_category_priority(classify_chapter_file(p)), p.name))

    rows: list[list[str]] = []
    quest_pairs: list[tuple[str, str, str]] = []

    for md_path in files:
        category = classify_chapter_file(md_path)
        if category == "npc":
            # 路人觸發矩陣統一進入總表，不放在各章整合頁重複出現
            continue

        sections = read_md_sections(md_path)
        flat = flatten_sections(sections)
        rows.append([f"## {section_title_for_file(md_path, category)}"])
        rows.extend(flat)
        rows.append([])
        quest_pairs.extend(extract_sequence_quest_pairs(flat, md_path.name))

    if not rows:
        rows = [["## EMPTY"], [f"找不到 {chapter_name}_*.md 可匯入內容"]]

    while rows and not any(str(c).strip() for c in rows[-1]):
        rows.pop()
    return normalize_rows(rows), quest_pairs


def build_npc_aggregate_rows(chapter_indices: Iterable[int]) -> list[list[str]]:
    rows: list[list[str]] = []
    for idx, chapter_name in CHAPTERS:
        if idx not in chapter_indices:
            continue
        path = WRITER_DIR / f"{chapter_name}{NPC_MATRIX_SUFFIX}"
        sections = read_md_sections(path)
        rows.append([f"## {chapter_name} | 路人對話觸發矩陣"])
        rows.extend(flatten_sections(sections))
        rows.append([])

    if not rows:
        rows = [["## EMPTY"], ["未選取任何章節，或沒有可匯入的路人觸發矩陣"]]
    while rows and not any(str(c).strip() for c in rows[-1]):
        rows.pop()
    return normalize_rows(rows)


def append_quest_index(rows: list[list[str]], quest_pairs: list[tuple[str, str, str]]) -> list[list[str]]:
    if not quest_pairs:
        return rows

    dedup = sorted(set(quest_pairs), key=lambda x: (normalize_seq_for_sort(x[0]), x[1], x[2]))
    rows = [r[:] for r in rows]
    rows.append([])
    rows.append(["## 任務索引（依劇本序號自動整理）"])
    rows.append(["序號", "任務ID", "來源檔案", "備註"])
    for seq, quest_id, source in dedup:
        rows.append([seq, quest_id, source, "可直接把任務編號寫在劇本序號旁，這裡會自動收斂"])
    return normalize_rows(rows)


def process_chapters(
    wb,
    chapter_indices: Iterable[int],
    with_expansion: bool,
) -> list[str]:
    created: list[str] = []
    old_style_sheet_names: set[str] = set()
    for _, name in CHAPTERS:
        old_style_sheet_names.update(
            {
                f"{name}劇本",
                sheet_title(name, "劇本"),
                sheet_title(name, "主線任務"),
                sheet_title(name, "支線任務"),
                sheet_title(name, "路人矩陣"),
                sheet_title(name, "整合"),
            }
        )
    old_style_sheet_names.add(NPC_AGGREGATE_SHEET)

    for old in old_style_sheet_names:
        if old in wb.sheetnames:
            del wb[old]

    for idx, chapter_name in CHAPTERS:
        if idx not in chapter_indices:
            continue

        chapter_rows, quest_pairs = build_chapter_rows_and_quests(chapter_name)
        chapter_rows = append_quest_index(chapter_rows, quest_pairs)

        title = sheet_title(chapter_name, "整合")
        if title in wb.sheetnames:
            del wb[title]
        ws = wb.create_sheet(title=title)
        style_sheet(ws, chapter_rows)
        created.append(title)

    npc_rows = build_npc_aggregate_rows(chapter_indices)
    if NPC_AGGREGATE_SHEET in wb.sheetnames:
        del wb[NPC_AGGREGATE_SHEET]
    ws = wb.create_sheet(title=NPC_AGGREGATE_SHEET)
    style_sheet(ws, npc_rows)
    created.append(NPC_AGGREGATE_SHEET)

    if with_expansion:
        for sheet_name, md_name, related in EXPANSIONS:
            if not set(chapter_indices) & related:
                continue
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            sections = read_md_sections(WRITER_DIR / md_name)
            rows = flatten_sections(sections)
            ws = wb.create_sheet(title=sheet_name)
            style_sheet(ws, rows)
            created.append(sheet_name)

    return created


def reorder_sheets(wb) -> None:
    desired: list[str] = []
    for _, chapter_name in CHAPTERS:
        desired.append(sheet_title(chapter_name, "整合"))
    desired.append(NPC_AGGREGATE_SHEET)
    desired.extend([x[0] for x in EXPANSIONS])

    for i, name in enumerate(desired):
        if name in wb.sheetnames:
            cur = wb.sheetnames.index(name)
            wb.move_sheet(name, offset=-(cur - i))


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    if args.list:
        print("可用章節：")
        for idx, chapter_name in CHAPTERS:
            print(f"  {idx}  {chapter_name}")
        print("\n匯入模式：")
        print("- 每章建立 1 張「整合」分頁（劇本/任務/補充合併）")
        print("- 任務編號會依劇本序號自動整理成索引")
        print(f"- 路人觸發對話彙整到單一分頁：{NPC_AGGREGATE_SHEET}")
        return

    selected = parse_chapter_args(args.chapters)
    chapter_indices = {idx for idx, _ in CHAPTERS} if selected is None else selected

    if not WORKBOOK_PATH.exists():
        raise SystemExit(f"[錯誤] 找不到 Excel：{WORKBOOK_PATH}")
    if not WRITER_DIR.exists():
        raise SystemExit(f"[錯誤] 找不到 writer 目錄：{WRITER_DIR}")

    print(f"讀取 Excel：{WORKBOOK_PATH}")
    wb = load_workbook(WORKBOOK_PATH)
    created = process_chapters(wb, chapter_indices, args.with_expansion)
    reorder_sheets(wb)

    if args.dry_run:
        print("DRY-RUN: 不寫入檔案")
    else:
        wb.save(WORKBOOK_PATH)
        print(f"已寫入：{WORKBOOK_PATH}")

    print("\n=== 匯入結果 ===")
    print(f"章節：{', '.join(str(x) for x in sorted(chapter_indices))}")
    print(f"建立/覆蓋分頁數：{len(created)}")
    for name in created:
        print(f"  - {name}")


if __name__ == "__main__":
    main()
