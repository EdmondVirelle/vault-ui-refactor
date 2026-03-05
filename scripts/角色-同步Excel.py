"""
將 Actors.json 的資料同步更新到 萬法同歸_資料庫.xlsx 的 Actors 工作表。
同步欄位：名稱、稱號、職業ID、初始等級、最大等級、個人簡介、備註(Note)
"""
import json
import openpyxl

ACTORS_PATH = r'C:\Consilience\Consilience\data\Actors.json'
XLSX_PATH = r'C:\Consilience\docs\萬法同歸_資料庫.xlsx'

# 欄位索引（1-based）
COL_ID = 1
COL_NAME = 2
COL_NICKNAME = 3
COL_CLASS_ID = 4
COL_INIT_LV = 5
COL_MAX_LV = 6
COL_CHAR_IMG = 7
COL_FACE_IMG = 8
COL_PROFILE = 9
COL_NOTE = 10


def main():
    # 讀取 Actors.json
    with open(ACTORS_PATH, 'r', encoding='utf-8') as f:
        actors = json.load(f)

    # 建立 ID → actor 的對應
    actor_map = {}
    for a in actors:
        if a is not None:
            actor_map[a['id']] = a

    # 讀取 XLSX
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb['Actors']

    updated = []

    for row_idx in range(2, ws.max_row + 1):
        cell_id = ws.cell(row=row_idx, column=COL_ID).value
        if cell_id is None:
            continue
        aid = int(cell_id)
        if aid not in actor_map:
            continue

        actor = actor_map[aid]

        # 同步名稱
        ws.cell(row=row_idx, column=COL_NAME).value = actor['name']
        # 同步稱號
        ws.cell(row=row_idx, column=COL_NICKNAME).value = actor.get('nickname', '')
        # 同步職業ID
        ws.cell(row=row_idx, column=COL_CLASS_ID).value = actor.get('classId', '')
        # 同步等級
        ws.cell(row=row_idx, column=COL_INIT_LV).value = actor.get('initialLevel', '')
        ws.cell(row=row_idx, column=COL_MAX_LV).value = actor.get('maxLevel', '')
        # 同步簡介
        ws.cell(row=row_idx, column=COL_PROFILE).value = actor.get('profile', '')
        # 同步備註（note）
        ws.cell(row=row_idx, column=COL_NOTE).value = actor.get('note', '')

        updated.append(f"  ID {aid:2d} {actor['name']}")

    wb.save(XLSX_PATH)
    print(f"已同步 {len(updated)} 位角色到 XLSX：")
    for line in updated:
        print(line)


if __name__ == '__main__':
    main()
