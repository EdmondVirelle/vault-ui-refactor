# -*- coding: utf-8 -*-
"""
萬法同歸 — 傷害計算表 v2 (全面更新版)
Sheets:
  1. 我方技能      — 角色技能 (desc, note, power, formula coefficients)
  2. 敵方技能      — 敵人技能 (desc, note, power, formula coefficients)
  3. 輕功總覽      — 輕功裝備 + 輕功技能合併表
  4. 角色數值      — 全角色 3 職業 × Lv20/50/99 完整參數
  5. 敵人數值      — 全敵人參數 + 元素弱點
  6. 裝備總覽      — 武器 + 防具完整資料
  7. 元素抗性      — 防禦端元素倍率矩陣
  8. 傷害計算機    — 互動式計算（含下拉選單與 Excel 公式）
  Hidden sheets:
  _角色參數      — 完整 level curve 供 VLOOKUP
  _裝備參數      — 扁平化裝備參數
  _技能係數      — 從公式抽取的攻/防/速/運係數
  _敵人參數      — 敵人 8 維 + 元素抗性
"""
import json, re, math, os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                             numbers, NamedStyle)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

os.environ['PYTHONUTF8'] = '1'

DATA_DIR = Path("C:/Consilience/Consilience/data")
OUT_PATH = Path("C:/Consilience/docs/萬法同歸_傷害計算表.xlsx")

# ── Load ─────────────────────────────────────────────────────────────
def load_json(name):
    with open(DATA_DIR / name, "r", encoding="utf-8") as f:
        return json.load(f)

skills_data  = load_json("Skills.json")
enemies_data = load_json("Enemies.json")
actors_data  = load_json("Actors.json")
classes_data = load_json("Classes.json")
weapons_data = load_json("Weapons.json")
armors_data  = load_json("Armors.json")
system_data  = load_json("System.json")

MAPPING_PATH = DATA_DIR / "actor_class_mapping.json"
if MAPPING_PATH.exists():
    with open(MAPPING_PATH, "r", encoding="utf-8") as f:
        ACTOR_CLASS_MAP = {int(k): v for k, v in json.load(f).items()}
else:
    ACTOR_CLASS_MAP = {}

# ── Constants ────────────────────────────────────────────────────────
PARAM_NAMES = ["氣血","內力","外功","外防","內功","內防","輕功","福緣"]
JOB_LABELS  = {0: "均衡", 1: "攻擊", 2: "防禦"}

raw_elements = system_data["elements"]
ELEMENTS = []
for e in raw_elements:
    c = re.sub(r'\\I\[\d+\]', '', e).strip()
    ELEMENTS.append(c if c else "(無)")
CORE_ELEM_IDS = list(range(1, 17))
CORE_ELEM_NAMES = [ELEMENTS[i] for i in CORE_ELEM_IDS]

ETYPE_NAMES = system_data.get("equipTypes", [])
WTYPE_NAMES = system_data.get("weaponTypes", [])

def strip_codes(text):
    if not text: return ''
    return re.sub(r'\\[cCiI]\[\d+\]', '', text).strip()

# ── Styles ───────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="2B3A4A", end_color="2B3A4A", fill_type="solid")
HEADER_FONT = Font(name="Microsoft JhengHei", size=10, bold=True, color="FFFFFF")
SUB_FILL    = PatternFill(start_color="3D5A6E", end_color="3D5A6E", fill_type="solid")
DATA_FONT   = Font(name="Microsoft JhengHei", size=9)
NUM_FONT    = Font(name="Consolas", size=9)
TITLE_FONT  = Font(name="Microsoft JhengHei", size=14, bold=True, color="2B3A4A")
LABEL_FONT  = Font(name="Microsoft JhengHei", size=10, bold=True, color="333333")
INPUT_FILL  = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
RESULT_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="C0C0C0"),
    right=Side(style="thin", color="C0C0C0"),
    top=Side(style="thin", color="C0C0C0"),
    bottom=Side(style="thin", color="C0C0C0"),
)
JOB_FILLS = {
    0: None,
    1: PatternFill(start_color="FFF5E6", end_color="FFF5E6", fill_type="solid"),
    2: PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid"),
}

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

def sc(cell, is_num=False, bold=False):
    """Style a data cell."""
    cell.font = NUM_FONT if is_num else DATA_FONT
    if bold:
        cell.font = Font(name=cell.font.name, size=cell.font.size, bold=True)
    cell.border = THIN_BORDER
    if is_num:
        cell.alignment = Alignment(horizontal="right")

# ── Element rates (code 11 = defensive) ──────────────────────────────
def get_elem_rates_defense(traits):
    """code 11 = Element Rate (defensive multiplier, >1 = weak, <1 = resist)."""
    rates = {}
    for t in traits:
        if t["code"] == 11:
            eid = t["dataId"]
            rates[eid] = rates.get(eid, 1.0) * t["value"]
    return rates

# ── Formula coefficient extraction ───────────────────────────────────
def extract_coefficients(formula):
    """Parse JS damage formula into coefficients for the Excel calculator."""
    c = {'a_atk': 0, 'b_def': 0, 'a_mat': 0, 'b_mdf': 0,
         'agi_div': 250, 'luk_coeff': 0, 'type': 'standard',
         'hp_ratio_mult': 0, 'pct_mhp': 0}

    if not formula or formula == "0":
        c['type'] = 'none'
        return c

    # Percent-based: b.mhp * X
    m = re.search(r'b\.mhp\s*\*\s*(\d+\.?\d*)', formula)
    if m and 'a.atk' not in formula and 'a.mat' not in formula:
        c['type'] = 'pct_mhp'
        c['pct_mhp'] = float(m.group(1))
        return c

    # HP-ratio based: a.hp/a.mhp * X
    m = re.search(r'a\.hp\s*/\s*a\.mhp\s*\*\s*(\d+\.?\d*)', formula)
    if m:
        c['type'] = 'hp_ratio'
        c['hp_ratio_mult'] = float(m.group(1))
        # Also extract b.def coeff
        m2 = re.search(r'b\.def\s*\*\s*(\d+\.?\d*)', formula)
        if m2: c['b_def'] = float(m2.group(1))
        m2 = re.search(r'a\.agi-b\.agi\)\s*/\s*(\d+)', formula)
        if m2: c['agi_div'] = int(m2.group(1))
        return c

    # Standard pattern: a.atk*X - b.def*Y + a.mat*W - b.mdf*Z
    m = re.search(r'a\.atk\s*\*\s*(\d+\.?\d*)', formula)
    if m: c['a_atk'] = float(m.group(1))

    m = re.search(r'b\.def\s*\*\s*(\d+\.?\d*)', formula)
    if m: c['b_def'] = float(m.group(1))

    m = re.search(r'a\.mat\s*\*\s*(\d+\.?\d*)', formula)
    if m: c['a_mat'] = float(m.group(1))

    m = re.search(r'b\.mdf\s*\*\s*(\d+\.?\d*)', formula)
    if m: c['b_mdf'] = float(m.group(1))

    m = re.search(r'a\.agi-b\.agi\)\s*/\s*(\d+)', formula)
    if m: c['agi_div'] = int(m.group(1))

    m = re.search(r'a\.luk-b\.luk\)\s*\*\s*(\d+\.?\d*)', formula)
    if m: c['luk_coeff'] = float(m.group(1))

    return c

# ── Damage formula evaluator ─────────────────────────────────────────
class FProxy:
    def __init__(self, stats_dict):
        for k, v in stats_dict.items():
            setattr(self, k, v)
        self.__dict__["def"] = stats_dict.get("def", stats_dict.get("df", 0))

def eval_formula(formula_str, a_stats, b_stats):
    if not formula_str or formula_str == "0":
        return 0
    a = FProxy(a_stats)
    b = FProxy(b_stats)
    f = formula_str
    f = f.replace("Math.max", "max").replace("Math.min", "min")
    f = f.replace("Math.abs", "abs").replace("Math.floor", "int")
    f = f.replace("Math.ceil", "math.ceil").replace("Math.round", "round")
    f = f.replace("Math.random()", "0.5").replace("Math.PI", "3.14159")
    f = re.sub(r'\ba\.def\b', 'a.__dict__["def"]', f)
    f = re.sub(r'\bb\.def\b', 'b.__dict__["def"]', f)
    f = f.replace("power", "1")
    f = re.sub(r'(\S+)\s*\?\s*(\S+)\s*:\s*(\S+)', r'(\2 if \1 else \3)', f)
    f = re.sub(r'v\[(\d+)\]', '0', f)
    f = re.sub(r'\$gameVariables\.value\(\d+\)', '0', f)
    f = f.replace("user.", "a.").replace("target.", "b.")
    try:
        result = eval(f, {"__builtins__": {}, "max": max, "min": min, "abs": abs,
                          "int": int, "round": round, "math": math, "a": a, "b": b})
        return max(0, round(result)) if isinstance(result, (int, float)) else 0
    except Exception:
        return -1

# ── Data builders ────────────────────────────────────────────────────
SCOPE_MAP = {0:'無', 1:'敵單', 2:'敵全', 3:'敵1隨機', 4:'敵2隨機',
             5:'敵3隨機', 6:'敵4隨機', 7:'我單', 8:'我全',
             9:'己方死亡單', 10:'己方死亡全', 11:'使用者',
             12:'敵單(即死)', 13:'我全(含死)', 14:'敵全(含死)'}
DTYPE_MAP = {0:'無', 1:'物理', 2:'魔法', 3:'必中', 4:'物理', 5:'吸收', 6:'自訂'}

def get_actor_info():
    """Return list of (actor_id, actor_name, [(job_index, class_id, job_name)])."""
    result = []
    for a in actors_data:
        if not a or not a.get("name", "").strip():
            continue
        aid = a["id"]
        mapping = ACTOR_CLASS_MAP.get(aid)
        if mapping:
            jobs = list(zip(range(3), mapping["classes"], mapping["job_names"]))
        else:
            cid = a["classId"]
            cls = classes_data[cid]
            jobs = [(0, cid, cls["name"] if cls else "")]
        result.append((aid, a["name"], jobs))
    return result

def get_stats_at_level(class_id, level):
    """Return 8 base stats for a class at given level."""
    cls = classes_data[class_id]
    if not cls:
        return [0]*8
    idx = min(level - 1, len(cls["params"][0]) - 1)
    return [cls["params"][p][idx] for p in range(8)]

def get_enemy_skills():
    """Return {enemy_id: [skill_ids]} from enemy actions."""
    result = {}
    for e in enemies_data:
        if not e or not e.get("name", "").strip() or e["name"].startswith("---"):
            continue
        sids = list(set(act["skillId"] for act in e.get("actions", [])))
        if sids:
            result[e["id"]] = sorted(sids)
    return result

def get_actor_skill_ids():
    """Return {actor_id: set(skill_ids)} from all class learnings."""
    result = {}
    for a in actors_data:
        if not a or not a.get("name", "").strip():
            continue
        aid = a["id"]
        sids = set()
        mapping = ACTOR_CLASS_MAP.get(aid)
        if mapping:
            for cid in mapping["classes"]:
                if cid < len(classes_data) and classes_data[cid]:
                    for l in classes_data[cid].get("learnings", []):
                        sids.add(l["skillId"])
        else:
            cid = a["classId"]
            if classes_data[cid]:
                for l in classes_data[cid].get("learnings", []):
                    sids.add(l["skillId"])
        result[aid] = sids
    return result

# ── Estimate skill "power" ───────────────────────────────────────────
def estimate_power(formula, elem_id=0):
    """Evaluate formula with standardized stats (Lv50 avg) to get base power."""
    ref_atk = {"atk": 50, "df": 50, "def": 50, "mat": 50, "mdf": 50,
               "agi": 50, "luk": 50, "hp": 500, "mhp": 500, "mp": 300, "mmp": 300}
    ref_def = {"atk": 40, "df": 40, "def": 40, "mat": 40, "mdf": 40,
               "agi": 40, "luk": 40, "hp": 800, "mhp": 800, "mp": 200, "mmp": 200}
    return eval_formula(formula, ref_atk, ref_def)


# ═══════════════════════════════════════════════════════════════════════
#  SHEET 1: 我方技能
# ═══════════════════════════════════════════════════════════════════════
def write_player_skills(wb):
    ws = wb.create_sheet("我方技能")
    headers = ["ID", "技能名稱", "說明", "範圍", "MP消耗", "TP消耗", "TP獲得",
               "傷害類型", "元素", "公式", "變異%", "連擊", "速度", "冷卻",
               "威力估算", "外功係數", "外防係數", "內功係數", "內防係數",
               "速度除數", "運勢係數", "效果", "備註(Note)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    actor_skills = get_actor_skill_ids()
    all_actor_sids = set()
    for sids in actor_skills.values():
        all_actor_sids |= sids

    row = 2
    for s in skills_data:
        if not s or not s.get("name", "").strip():
            continue
        if s["name"].startswith("---") or s["name"].startswith("__"):
            continue
        sid = s["id"]
        # Include: skills in class learnings, OR stypeId > 0, OR 輕功 skills
        if sid not in all_actor_sids and s.get("stypeId", 0) == 0 and not (601 <= sid <= 638):
            continue

        dmg = s.get("damage", {})
        formula = dmg.get("formula", "")
        coeffs = extract_coefficients(formula)
        desc = strip_codes(s.get("description", ""))
        note = s.get("note", "")

        # Extract cooldown from note
        cd_match = re.search(r'<Cooldown:\s*(\d+)>', note)
        cooldown = int(cd_match.group(1)) if cd_match else 0

        effects = s.get("effects", [])
        eff_str = "; ".join(
            f"code={e['code']} id={e.get('dataId','')} v1={e.get('value1','')} v2={e.get('value2','')}"
            for e in effects
        ) if effects else ""

        elem_id = dmg.get("elementId", 0)
        elem_name = ELEMENTS[elem_id] if 0 < elem_id < len(ELEMENTS) else "(無)"
        dtype = dmg.get("type", 0)
        power = estimate_power(formula, elem_id) if dtype > 0 and formula and formula != "0" else 0

        vals = [
            sid, s["name"], desc,
            SCOPE_MAP.get(s.get("scope", 0), str(s.get("scope", 0))),
            s.get("mpCost", 0), s.get("tpCost", 0), s.get("tpGain", 0),
            DTYPE_MAP.get(dtype, str(dtype)), elem_name, formula,
            dmg.get("variance", 20), s.get("repeats", 1),
            s.get("speed", 0), cooldown, power,
            coeffs['a_atk'], coeffs['b_def'], coeffs['a_mat'], coeffs['b_mdf'],
            coeffs['agi_div'], coeffs['luk_coeff'],
            eff_str, note
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            sc(cell, ci in (1, 5, 6, 7, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21))
        row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row-1}"
    ws.freeze_panes = "C2"
    widths = [5, 16, 35, 8, 7, 7, 7, 7, 7, 60, 6, 5, 6, 5, 8,
              7, 7, 7, 7, 7, 7, 30, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return row - 2


# ═══════════════════════════════════════════════════════════════════════
#  SHEET 2: 敵方技能
# ═══════════════════════════════════════════════════════════════════════
def write_enemy_skills(wb):
    ws = wb.create_sheet("敵方技能")
    headers = ["技能ID", "技能名稱", "說明", "範圍", "MP消耗",
               "傷害類型", "元素", "公式", "變異%", "連擊", "速度",
               "威力估算", "外功係數", "外防係數", "內功係數", "內防係數",
               "速度除數", "運勢係數", "使用敵人", "效果", "備註(Note)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    enemy_skill_map = get_enemy_skills()
    # Reverse: skill_id → [enemy_names]
    skill_enemies = {}
    for eid, sids in enemy_skill_map.items():
        e = enemies_data[eid] if eid < len(enemies_data) else None
        if not e:
            continue
        for sid in sids:
            skill_enemies.setdefault(sid, []).append(e["name"])

    row = 2
    seen = set()
    for eid, sids in sorted(enemy_skill_map.items()):
        for sid in sids:
            if sid in seen:
                continue
            seen.add(sid)
            if sid >= len(skills_data) or not skills_data[sid]:
                continue
            s = skills_data[sid]
            if not s.get("name", "").strip():
                continue

            dmg = s.get("damage", {})
            formula = dmg.get("formula", "")
            coeffs = extract_coefficients(formula)
            desc = strip_codes(s.get("description", ""))
            note = s.get("note", "")

            elem_id = dmg.get("elementId", 0)
            elem_name = ELEMENTS[elem_id] if 0 < elem_id < len(ELEMENTS) else "(無)"
            dtype = dmg.get("type", 0)
            power = estimate_power(formula, elem_id) if dtype > 0 and formula and formula != "0" else 0

            users = ", ".join(skill_enemies.get(sid, []))

            effects = s.get("effects", [])
            eff_str = "; ".join(
                f"code={e['code']} id={e.get('dataId','')} v1={e.get('value1','')} v2={e.get('value2','')}"
                for e in effects
            ) if effects else ""

            vals = [
                sid, s["name"], desc,
                SCOPE_MAP.get(s.get("scope", 0), str(s.get("scope", 0))),
                s.get("mpCost", 0),
                DTYPE_MAP.get(dtype, str(dtype)), elem_name, formula,
                dmg.get("variance", 20), s.get("repeats", 1),
                s.get("speed", 0), power,
                coeffs['a_atk'], coeffs['b_def'], coeffs['a_mat'], coeffs['b_mdf'],
                coeffs['agi_div'], coeffs['luk_coeff'],
                users, eff_str, note
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=ci, value=v)
                sc(cell, ci in (1, 5, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18))
            row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row-1}"
    ws.freeze_panes = "C2"
    widths = [6, 16, 35, 8, 7, 7, 7, 60, 6, 5, 6, 8, 7, 7, 7, 7, 7, 7, 25, 30, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return row - 2


# ═══════════════════════════════════════════════════════════════════════
#  SHEET 3: 輕功總覽
# ═══════════════════════════════════════════════════════════════════════
def write_lightfoot(wb):
    ws = wb.create_sheet("輕功總覽")
    headers = ["裝備ID", "輕功名稱", "等級", "輕功加成(AGI)", "技能ID",
               "技能效果", "技能說明", "冷卻", "備註(Note)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    row = 2
    for a in armors_data:
        if not a or a.get("etypeId") != 5 or not a.get("name", "").strip():
            continue
        aid = a["id"]
        agi_bonus = a["params"][6]
        desc = strip_codes(a.get("description", ""))

        # Extract linked skill ID from trait code 43
        skill_id = 0
        for t in a.get("traits", []):
            if t["code"] == 43:
                skill_id = t["dataId"]

        # Get skill data
        skill_effect = ""
        skill_desc = ""
        cooldown = 0
        note = ""
        if skill_id and skill_id < len(skills_data) and skills_data[skill_id]:
            sk = skills_data[skill_id]
            skill_desc = strip_codes(sk.get("description", ""))
            snote = sk.get("note", "")
            note = snote

            # Parse OTB effects from note
            effects = []
            for line in snote.split("\n"):
                line = line.strip()
                if line.startswith("<OTB"):
                    effects.append(line)
                elif line.startswith("<Cooldown"):
                    m = re.search(r'<Cooldown:\s*(\d+)>', line)
                    if m: cooldown = int(m.group(1))
            skill_effect = " | ".join(effects)

        # Extract 輕功 level from desc
        lv_match = re.search(r'輕功[：:](\d+)', desc)
        lv = int(lv_match.group(1)) if lv_match else agi_bonus

        vals = [aid, a["name"], lv, agi_bonus, skill_id,
                skill_effect, skill_desc, cooldown, note]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            sc(cell, ci in (1, 3, 4, 5, 8))
            # Color code by tier
            if lv <= 5:
                pass  # white
            elif lv <= 20:
                cell.fill = PatternFill(start_color="FFF8E1", fill_type="solid")
            elif lv <= 50:
                cell.fill = PatternFill(start_color="FFE0B2", fill_type="solid")
            elif lv <= 80:
                cell.fill = PatternFill(start_color="FFCCBC", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFCDD2", fill_type="solid")
        row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row-1}"
    ws.freeze_panes = "C2"
    widths = [7, 14, 6, 12, 7, 40, 35, 6, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return row - 2


# ═══════════════════════════════════════════════════════════════════════
#  SHEET 4: 角色數值
# ═══════════════════════════════════════════════════════════════════════
def write_actor_stats(wb):
    ws = wb.create_sheet("角色數值")
    headers = ["ID", "角色", "職業", "類型", "等級"] + PARAM_NAMES
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    actors = get_actor_info()
    row = 2
    for level in [20, 50, 99]:
        for aid, aname, jobs in actors:
            for ji, cid, jname in jobs:
                stats = get_stats_at_level(cid, level)
                vals = [aid, aname, jname, JOB_LABELS.get(ji, ""), level] + stats
                for ci, v in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=ci, value=v)
                    sc(cell, ci >= 5)
                    jf = JOB_FILLS.get(ji)
                    if jf and ci <= 4:
                        cell.fill = jf
                row += 1

    ws.auto_filter.ref = f"A1:M{row-1}"
    ws.freeze_panes = "F2"
    widths = [5, 10, 12, 6, 6] + [8]*8
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return row - 2


# ═══════════════════════════════════════════════════════════════════════
#  SHEET 5: 敵人數值
# ═══════════════════════════════════════════════════════════════════════
def write_enemy_stats(wb):
    ws = wb.create_sheet("敵人數值")
    headers = ["ID", "敵人", "經驗", "金幣"] + PARAM_NAMES + CORE_ELEM_NAMES + ["備註(Note)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    row = 2
    for e in enemies_data:
        if not e or not e.get("name", "").strip() or e["name"].startswith("---"):
            continue
        p = e["params"]
        elem_rates = get_elem_rates_defense(e.get("traits", []))

        vals = [e["id"], e["name"], e.get("exp", 0), e.get("gold", 0)]
        vals += p
        for eid in CORE_ELEM_IDS:
            vals.append(elem_rates.get(eid, 1.0))
        vals.append(e.get("note", "")[:200])

        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            sc(cell, ci >= 3)
            # Color element rate cells
            if 13 <= ci <= 28:
                if isinstance(v, (int, float)):
                    cell.number_format = '0%'
                    if v > 1.0:
                        cell.fill = PatternFill(start_color="FFC0C0", fill_type="solid")
                    elif v < 1.0:
                        cell.fill = PatternFill(start_color="C0FFC0", fill_type="solid")
        row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row-1}"
    ws.freeze_panes = "C2"
    widths = [5, 14, 7, 7] + [8]*8 + [7]*16 + [30]
    for i, w in enumerate(widths[:len(headers)], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return row - 2


# ═══════════════════════════════════════════════════════════════════════
#  SHEET 6: 裝備總覽
# ═══════════════════════════════════════════════════════════════════════
def write_equipment(wb):
    ws = wb.create_sheet("裝備總覽")
    headers = ["ID", "名稱", "類別", "子類型", "說明", "價格"] + PARAM_NAMES + ["特性", "備註(Note)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    row = 2
    # Weapons first
    for w in weapons_data:
        if not w or not w.get("name", "").strip():
            continue
        wt = w.get("wtypeId", 0)
        wt_name = strip_codes(WTYPE_NAMES[wt]) if wt < len(WTYPE_NAMES) else str(wt)
        p = w.get("params", [0]*8)
        traits = w.get("traits", [])
        trait_str = "; ".join(
            f"c={t['code']}d={t.get('dataId','')}v={t.get('value','')}"
            for t in traits
        ) if traits else ""

        vals = [w["id"], w["name"], "武器", wt_name,
                strip_codes(w.get("description", "")),
                w.get("price", 0)] + list(p) + [trait_str, w.get("note", "")]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            sc(cell, 6 <= ci <= 14)
        row += 1

    # Armors
    for a in armors_data:
        if not a or not a.get("name", "").strip():
            continue
        et = a.get("etypeId", 0)
        et_name = strip_codes(ETYPE_NAMES[et]) if et < len(ETYPE_NAMES) else str(et)
        p = a.get("params", [0]*8)
        at = a.get("atypeId", 0)
        traits = a.get("traits", [])
        trait_str = "; ".join(
            f"c={t['code']}d={t.get('dataId','')}v={t.get('value','')}"
            for t in traits
        ) if traits else ""

        vals = [a["id"], a["name"], et_name, str(at),
                strip_codes(a.get("description", "")),
                a.get("price", 0)] + list(p) + [trait_str, a.get("note", "")]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            sc(cell, 6 <= ci <= 14)
        row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row-1}"
    ws.freeze_panes = "C2"
    widths = [6, 14, 7, 8, 30, 8] + [8]*8 + [25, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return row - 2


# ═══════════════════════════════════════════════════════════════════════
#  SHEET 7: 元素抗性
# ═══════════════════════════════════════════════════════════════════════
def write_element_rates(wb):
    ws = wb.create_sheet("元素抗性")
    headers = ["類型", "ID", "名稱"] + CORE_ELEM_NAMES
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    row = 2
    # Enemies
    for e in enemies_data:
        if not e or not e.get("name", "").strip() or e["name"].startswith("---"):
            continue
        rates = get_elem_rates_defense(e.get("traits", []))
        ws.cell(row=row, column=1, value="敵人")
        ws.cell(row=row, column=2, value=e["id"])
        ws.cell(row=row, column=3, value=e["name"])
        for ci, eid in enumerate(CORE_ELEM_IDS, 4):
            rate = rates.get(eid, 1.0)
            cell = ws.cell(row=row, column=ci, value=rate)
            cell.number_format = '0%'
            sc(cell, True)
            if rate > 1.0:
                cell.fill = PatternFill(start_color="FFC0C0", fill_type="solid")
            elif rate < 1.0:
                cell.fill = PatternFill(start_color="C0FFC0", fill_type="solid")
        for col in range(1, 4):
            sc(ws.cell(row=row, column=col))
        row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row-1}"
    ws.freeze_panes = "D2"
    widths = [6, 5, 14] + [7]*16
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return row - 2


# ═══════════════════════════════════════════════════════════════════════
#  HIDDEN SHEET: _角色參數 (full level curve for VLOOKUP)
# ═══════════════════════════════════════════════════════════════════════
def write_actor_params_lookup(wb):
    ws = wb.create_sheet("_角色參數")
    headers = ["查找鍵", "角色ID", "角色", "職業", "職業ID", "類型", "等級",
               "氣血", "內力", "外功", "外防", "內功", "內防", "輕功", "福緣"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)

    actors = get_actor_info()
    row = 2
    for aid, aname, jobs in actors:
        for ji, cid, jname in jobs:
            cls = classes_data[cid]
            if not cls:
                continue
            for lv in range(1, 100):
                idx = min(lv - 1, len(cls["params"][0]) - 1)
                stats = [cls["params"][p][idx] for p in range(8)]
                # lookup key: "角色名_職業_等級"
                key = f"{aname}_{jname}_{lv}"
                vals = [key, aid, aname, jname, cid, JOB_LABELS.get(ji, ""), lv] + stats
                for ci, v in enumerate(vals, 1):
                    ws.cell(row=row, column=ci, value=v)
                row += 1

    ws.sheet_state = 'hidden'
    return row - 2


# ═══════════════════════════════════════════════════════════════════════
#  HIDDEN SHEET: _裝備參數
# ═══════════════════════════════════════════════════════════════════════
def write_equip_lookup(wb):
    ws = wb.create_sheet("_裝備參數")
    headers = ["查找鍵", "ID", "名稱", "類別", "子類型",
               "氣血", "內力", "外功", "外防", "內功", "內防", "輕功", "福緣"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)

    row = 2
    # Weapons
    for w in weapons_data:
        if not w or not w.get("name", "").strip():
            continue
        wt = w.get("wtypeId", 0)
        wt_name = strip_codes(WTYPE_NAMES[wt]) if wt < len(WTYPE_NAMES) else str(wt)
        p = w.get("params", [0]*8)
        key = f"武器_{w['name']}"
        vals = [key, w["id"], w["name"], "武器", wt_name] + list(p)
        for ci, v in enumerate(vals, 1):
            ws.cell(row=row, column=ci, value=v)
        row += 1

    # Armors
    for a in armors_data:
        if not a or not a.get("name", "").strip():
            continue
        et = a.get("etypeId", 0)
        et_name = strip_codes(ETYPE_NAMES[et]) if et < len(ETYPE_NAMES) else str(et)
        p = a.get("params", [0]*8)
        key = f"{et_name}_{a['name']}"
        vals = [key, a["id"], a["name"], et_name, str(a.get("atypeId", 0))] + list(p)
        for ci, v in enumerate(vals, 1):
            ws.cell(row=row, column=ci, value=v)
        row += 1

    ws.sheet_state = 'hidden'
    return row - 2


# ═══════════════════════════════════════════════════════════════════════
#  HIDDEN SHEET: _敵人參數
# ═══════════════════════════════════════════════════════════════════════
def write_enemy_lookup(wb):
    ws = wb.create_sheet("_敵人參數")
    headers = ["敵人名稱", "ID", "氣血", "內力", "外功", "外防", "內功", "內防", "輕功", "福緣"]
    headers += [f"元素{eid}抗" for eid in CORE_ELEM_IDS]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)

    row = 2
    for e in enemies_data:
        if not e or not e.get("name", "").strip() or e["name"].startswith("---"):
            continue
        p = e["params"]
        rates = get_elem_rates_defense(e.get("traits", []))
        vals = [e["name"], e["id"]] + list(p)
        for eid in CORE_ELEM_IDS:
            vals.append(rates.get(eid, 1.0))
        for ci, v in enumerate(vals, 1):
            ws.cell(row=row, column=ci, value=v)
        row += 1

    ws.sheet_state = 'hidden'
    return row - 2


# ═══════════════════════════════════════════════════════════════════════
#  HIDDEN SHEET: _技能係數
# ═══════════════════════════════════════════════════════════════════════
def write_skill_coeff_lookup(wb):
    ws = wb.create_sheet("_技能係數")
    headers = ["技能名稱", "ID", "元素ID", "元素", "傷害類型",
               "外功係數", "外防係數", "內功係數", "內防係數",
               "速度除數", "運勢係數", "公式類型", "百分比值", "HP倍率",
               "公式", "變異%", "連擊"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)

    row = 2
    for s in skills_data:
        if not s or not s.get("name", "").strip():
            continue
        if s["name"].startswith("---") or s["name"].startswith("__"):
            continue
        dmg = s.get("damage", {})
        dtype = dmg.get("type", 0)
        formula = dmg.get("formula", "")
        if dtype <= 0:
            continue

        coeffs = extract_coefficients(formula)
        elem_id = dmg.get("elementId", 0)
        if elem_id < 0: elem_id = 0
        elem_name = ELEMENTS[elem_id] if 0 < elem_id < len(ELEMENTS) else "(無)"

        vals = [
            s["name"], s["id"], elem_id, elem_name,
            DTYPE_MAP.get(dtype, str(dtype)),
            coeffs['a_atk'], coeffs['b_def'], coeffs['a_mat'], coeffs['b_mdf'],
            coeffs['agi_div'], coeffs['luk_coeff'], coeffs['type'],
            coeffs['pct_mhp'], coeffs['hp_ratio_mult'],
            formula, dmg.get("variance", 20), s.get("repeats", 1)
        ]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=row, column=ci, value=v)
        row += 1

    ws.sheet_state = 'hidden'
    return row - 2


# ═══════════════════════════════════════════════════════════════════════
#  SHEET 8: 傷害計算機
# ═══════════════════════════════════════════════════════════════════════
def write_calculator(wb):
    ws = wb.create_sheet("傷害計算機")

    # -- Collect dropdown lists --
    actors_info = get_actor_info()
    actor_names = [a[1] for a in actors_info]
    all_job_names = []
    for _, _, jobs in actors_info:
        for _, _, jname in jobs:
            if jname not in all_job_names:
                all_job_names.append(jname)

    weapon_names = [f"武器_{w['name']}" for w in weapons_data
                    if w and w.get("name", "").strip()]
    body_names = [f"身體_{a['name']}" for a in armors_data
                  if a and a.get("etypeId") == 2 and a.get("name", "").strip()]
    ring_names = [f"指環_{a['name']}" for a in armors_data
                  if a and a.get("etypeId") == 3 and a.get("name", "").strip()]
    skill_names_list = [f"功法_{a['name']}" for a in armors_data
                        if a and a.get("etypeId") == 4 and a.get("name", "").strip()]
    lf_names = [f"輕功_{a['name']}" for a in armors_data
                if a and a.get("etypeId") == 5 and a.get("name", "").strip()]

    skill_dmg_names = []
    for s in skills_data:
        if not s or not s.get("name", "").strip():
            continue
        if s["name"].startswith("---") or s["name"].startswith("__"):
            continue
        dmg = s.get("damage", {})
        if dmg.get("type", 0) > 0 and dmg.get("formula", "") not in ("", "0"):
            skill_dmg_names.append(s["name"])

    enemy_names_list = []
    for e in enemies_data:
        if e and e.get("name", "").strip() and not e["name"].startswith("---"):
            enemy_names_list.append(e["name"])

    # -- Helper: define data validation --
    def add_dv(ws, cell_ref, items, max_items=200):
        """Add dropdown data validation. For long lists, use a formula ref."""
        if len(items) > max_items:
            items = items[:max_items]
        formula = '"' + ','.join(items[:200]) + '"'
        if len(formula) > 255:
            # Too long for inline list, truncate
            while len(formula) > 255:
                items = items[:-1]
                formula = '"' + ','.join(items) + '"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "請從下拉選單選擇"
        dv.errorTitle = "無效輸入"
        ws.add_data_validation(dv)
        dv.add(ws[cell_ref])

    # -- Lookup sheet references --
    actor_params_sheet = "_角色參數"
    equip_sheet = "_裝備參數"
    enemy_sheet = "_敵人參數"
    skill_sheet = "_技能係數"

    # -- Layout --
    R = 1  # current row

    # Title
    ws.cell(row=R, column=1, value="萬法同歸 — 傷害計算機")
    ws.cell(row=R, column=1).font = TITLE_FONT
    ws.merge_cells('A1:G1')
    R += 2

    # ── Input section ─────────────────────────────────────────────
    ws.cell(row=R, column=1, value="【輸入區】").font = LABEL_FONT
    R += 1

    labels = ["角色名稱", "職業名稱", "等級", "武器", "身體", "指環", "功法", "輕功",
              "技能名稱", "敵人名稱"]
    defaults = [actor_names[0] if actor_names else "",
                all_job_names[0] if all_job_names else "",
                50,
                weapon_names[0] if weapon_names else "",
                body_names[0] if body_names else "",
                ring_names[0] if ring_names else "",
                skill_names_list[0] if skill_names_list else "",
                lf_names[0] if lf_names else "",
                skill_dmg_names[0] if skill_dmg_names else "",
                enemy_names_list[0] if enemy_names_list else ""]

    input_start_row = R
    for i, (label, default) in enumerate(zip(labels, defaults)):
        r = R + i
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        cell = ws.cell(row=r, column=2, value=default)
        cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
        cell.font = DATA_FONT
        if label == "等級":
            dv = DataValidation(type="whole", operator="between",
                                formula1="1", formula2="99")
            ws.add_data_validation(dv)
            dv.add(ws.cell(row=r, column=2))

    # Add dropdown validations
    r_actor = input_start_row
    r_job = input_start_row + 1
    r_level = input_start_row + 2
    r_weapon = input_start_row + 3
    r_body = input_start_row + 4
    r_ring = input_start_row + 5
    r_skill_eq = input_start_row + 6
    r_lf = input_start_row + 7
    r_skill = input_start_row + 8
    r_enemy = input_start_row + 9

    add_dv(ws, f"B{r_actor}", actor_names)
    add_dv(ws, f"B{r_job}", all_job_names)
    add_dv(ws, f"B{r_weapon}", weapon_names[:150])
    add_dv(ws, f"B{r_body}", body_names[:150])
    add_dv(ws, f"B{r_ring}", ring_names[:100])
    add_dv(ws, f"B{r_skill_eq}", skill_names_list[:150])
    add_dv(ws, f"B{r_lf}", lf_names[:100])
    add_dv(ws, f"B{r_skill}", skill_dmg_names[:200])
    add_dv(ws, f"B{r_enemy}", enemy_names_list[:200])

    R += len(labels) + 1

    # ── Actor base stats (Excel formulas with VLOOKUP) ─────────────
    ws.cell(row=R, column=1, value="【角色基礎數值】").font = LABEL_FONT
    R += 1
    # Lookup key: actor_name & "_" & job_name & "_" & level
    lookup_key_cell = f'B{r_actor}&"_"&B{r_job}&"_"&B{r_level}'

    stat_labels = PARAM_NAMES
    base_stat_row = R
    for i, label in enumerate(stat_labels):
        r = R + i
        ws.cell(row=r, column=1, value=f"基礎{label}").font = DATA_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        # VLOOKUP into _角色參數 sheet: column 8+i (MHP=8, MMP=9, ATK=10, ...)
        col_idx = 8 + i
        formula = f'=IFERROR(VLOOKUP({lookup_key_cell},\'{actor_params_sheet}\'!A:O,{col_idx},FALSE),0)'
        cell = ws.cell(row=r, column=2, value=formula if isinstance(formula, str) else 0)
        cell.font = NUM_FONT
        cell.border = THIN_BORDER
    R += len(stat_labels) + 1

    # ── Equipment bonuses ──────────────────────────────────────────
    ws.cell(row=R, column=1, value="【裝備加成】").font = LABEL_FONT
    R += 1

    equip_rows = [
        ("武器", r_weapon),
        ("身體", r_body),
        ("指環", r_ring),
        ("功法", r_skill_eq),
        ("輕功", r_lf),
    ]
    equip_bonus_rows = {}  # param_index → row of total bonus
    equip_start = R

    # For each equipment slot, show the 8 stat bonuses
    for slot_label, slot_row in equip_rows:
        ws.cell(row=R, column=1, value=slot_label).font = DATA_FONT
        ws.cell(row=R, column=1).border = THIN_BORDER
        for pi in range(8):
            col = 2 + pi
            col_idx = 6 + pi  # params start at column 6 in _裝備參數
            formula = f'=IFERROR(VLOOKUP(B{slot_row},\'_裝備參數\'!A:M,{col_idx},FALSE),0)'
            cell = ws.cell(row=R, column=col, value=formula)
            cell.font = NUM_FONT
            cell.border = THIN_BORDER
        R += 1

    # Totals row
    ws.cell(row=R, column=1, value="裝備合計").font = LABEL_FONT
    ws.cell(row=R, column=1).border = THIN_BORDER
    equip_total_row = R
    for pi in range(8):
        col = 2 + pi
        sum_cells = "+".join(f"{get_column_letter(col)}{equip_start + i}" for i in range(len(equip_rows)))
        cell = ws.cell(row=R, column=col, value=f"={sum_cells}")
        cell.font = Font(name="Consolas", size=9, bold=True)
        cell.border = THIN_BORDER
    R += 1

    # Header labels for param columns
    for pi, pname in enumerate(PARAM_NAMES):
        ws.cell(row=equip_start - 1, column=2 + pi, value=pname).font = Font(
            name="Microsoft JhengHei", size=8, color="666666")

    R += 1

    # ── Final stats ────────────────────────────────────────────────
    ws.cell(row=R, column=1, value="【最終數值】").font = LABEL_FONT
    R += 1
    final_stat_row = R
    for pi, pname in enumerate(PARAM_NAMES):
        ws.cell(row=R, column=1, value=f"最終{pname}").font = DATA_FONT
        ws.cell(row=R, column=1).border = THIN_BORDER
        base_cell = f"B{base_stat_row + pi}"
        bonus_cell = f"{get_column_letter(2 + pi)}{equip_total_row}"
        cell = ws.cell(row=R, column=2, value=f"={base_cell}+{bonus_cell}")
        cell.font = Font(name="Consolas", size=10, bold=True)
        cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        cell.border = THIN_BORDER
        R += 1
    R += 1

    # Store final stat cell references
    f_mhp = f"B{final_stat_row}"
    f_mmp = f"B{final_stat_row+1}"
    f_atk = f"B{final_stat_row+2}"
    f_def = f"B{final_stat_row+3}"
    f_mat = f"B{final_stat_row+4}"
    f_mdf = f"B{final_stat_row+5}"
    f_agi = f"B{final_stat_row+6}"
    f_luk = f"B{final_stat_row+7}"

    # ── Enemy stats ────────────────────────────────────────────────
    ws.cell(row=R, column=1, value="【敵人數值】").font = LABEL_FONT
    R += 1
    enemy_stat_row = R
    for pi, pname in enumerate(PARAM_NAMES):
        ws.cell(row=R, column=1, value=f"敵{pname}").font = DATA_FONT
        ws.cell(row=R, column=1).border = THIN_BORDER
        col_idx = 3 + pi  # params start at column 3 in _敵人參數
        formula = f'=IFERROR(VLOOKUP(B{r_enemy},\'_敵人參數\'!A:Z,{col_idx},FALSE),0)'
        cell = ws.cell(row=R, column=2, value=formula)
        cell.font = NUM_FONT
        cell.border = THIN_BORDER
        R += 1

    e_mhp = f"B{enemy_stat_row}"
    e_mmp = f"B{enemy_stat_row+1}"
    e_atk = f"B{enemy_stat_row+2}"
    e_def = f"B{enemy_stat_row+3}"
    e_mat = f"B{enemy_stat_row+4}"
    e_mdf = f"B{enemy_stat_row+5}"
    e_agi = f"B{enemy_stat_row+6}"
    e_luk = f"B{enemy_stat_row+7}"
    R += 1

    # ── Skill coefficients ─────────────────────────────────────────
    ws.cell(row=R, column=1, value="【技能資訊】").font = LABEL_FONT
    R += 1
    skill_info_row = R

    skill_fields = [
        ("技能元素ID", 3),
        ("技能元素", 4),
        ("傷害類型", 5),
        ("外功係數(a)", 6),
        ("外防係數(b)", 7),
        ("內功係數(a)", 8),
        ("內防係數(b)", 9),
        ("速度除數", 10),
        ("運勢係數", 11),
        ("公式類型", 12),
        ("百分比值", 13),
        ("HP倍率", 14),
        ("公式", 15),
        ("變異%", 16),
        ("連擊數", 17),
    ]
    for i, (label, col_idx) in enumerate(skill_fields):
        r = R + i
        ws.cell(row=r, column=1, value=label).font = DATA_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        formula = f'=IFERROR(VLOOKUP(B{r_skill},\'_技能係數\'!A:Q,{col_idx},FALSE),"")'
        cell = ws.cell(row=r, column=2, value=formula)
        cell.font = NUM_FONT
        cell.border = THIN_BORDER
    R += len(skill_fields) + 1

    # Coeff cell references
    s_a_atk = f"B{skill_info_row + 5}"  # 外功係數
    s_b_def = f"B{skill_info_row + 6}"  # 外防係數
    s_a_mat = f"B{skill_info_row + 7}"  # 內功係數
    s_b_mdf = f"B{skill_info_row + 8}"  # 內防係數
    s_agi_div = f"B{skill_info_row + 9}"  # 速度除數
    s_luk_coeff = f"B{skill_info_row + 10}"  # 運勢係數
    s_type = f"B{skill_info_row + 11}"  # 公式類型
    s_pct = f"B{skill_info_row + 12}"  # 百分比值
    s_hp_mult = f"B{skill_info_row + 13}"  # HP倍率
    s_variance = f"B{skill_info_row + 15}"  # 變異
    s_repeats = f"B{skill_info_row + 16}"  # 連擊
    s_elem_id = f"B{skill_info_row}"  # 元素ID

    # ── Element rate for this skill vs this enemy ──────────────────
    ws.cell(row=R, column=1, value="【元素倍率】").font = LABEL_FONT
    R += 1
    elem_rate_row = R
    ws.cell(row=R, column=1, value="敵人元素抗性").font = DATA_FONT
    ws.cell(row=R, column=1).border = THIN_BORDER
    # Lookup: in _敵人參數, element rates start at column 11 (after 2+8 = 10)
    # elem_id is 1-16, column is 10 + elem_id
    formula = (
        f'=IF({s_elem_id}=0, 1, '
        f'IF({s_elem_id}<0, 1, '
        f'IFERROR(INDEX(\'_敵人參數\'!A:Z, '
        f'MATCH(B{r_enemy},\'_敵人參數\'!A:A,0), '
        f'10+{s_elem_id}), 1)))'
    )
    cell = ws.cell(row=R, column=2, value=formula)
    cell.font = NUM_FONT
    cell.number_format = '0%'
    cell.border = THIN_BORDER
    R += 2

    # ── DAMAGE CALCULATION ──────────────────────────────────────────
    ws.cell(row=R, column=1, value="【傷害計算結果】").font = Font(
        name="Microsoft JhengHei", size=12, bold=True, color="C62828")
    R += 1

    # Base damage (standard formula)
    dmg_calc_row = R
    ws.cell(row=R, column=1, value="基礎傷害(公式)").font = LABEL_FONT
    ws.cell(row=R, column=1).border = THIN_BORDER
    # Standard: MAX(1, (ATK*a_atk - DEF*b_def + MAT*a_mat - MDF*b_mdf) * (1+(AGI-E_AGI)/agi_div) + (LUK-E_LUK)*luk_coeff)
    # Percent: E_MHP * pct
    # HP ratio: (HP/MHP * hp_mult - E_DEF*b_def) * (1+(AGI-E_AGI)/agi_div)
    formula_std = (
        f'=IF({s_type}="pct_mhp", {e_mhp}*{s_pct}, '
        f'IF({s_type}="hp_ratio", '
        f'MAX(1, ({f_mhp}/{f_mhp}*{s_hp_mult}-{e_def}*{s_b_def})*'
        f'(1+({f_agi}-{e_agi})/IF({s_agi_div}=0,250,{s_agi_div}))), '
        f'MAX(1, ({f_atk}*{s_a_atk}-{e_def}*{s_b_def}+{f_mat}*{s_a_mat}-{e_mdf}*{s_b_mdf})*'
        f'(1+({f_agi}-{e_agi})/IF({s_agi_div}=0,250,{s_agi_div}))'
        f'+MAX(0,({f_luk}-{e_luk})*{s_luk_coeff}))))'
    )
    cell = ws.cell(row=R, column=2, value=formula_std)
    cell.font = NUM_FONT
    cell.border = THIN_BORDER
    base_dmg = f"B{R}"
    R += 1

    # Element rate applied
    ws.cell(row=R, column=1, value="元素倍率").font = LABEL_FONT
    ws.cell(row=R, column=1).border = THIN_BORDER
    cell = ws.cell(row=R, column=2, value=f"=B{elem_rate_row}")
    cell.font = NUM_FONT
    cell.number_format = '0%'
    cell.border = THIN_BORDER
    elem_cell = f"B{R}"
    R += 1

    # Final damage per hit
    ws.cell(row=R, column=1, value="每擊傷害").font = LABEL_FONT
    ws.cell(row=R, column=1).border = THIN_BORDER
    cell = ws.cell(row=R, column=2, value=f"=ROUND(MAX(0,{base_dmg}*{elem_cell}),0)")
    cell.font = Font(name="Consolas", size=12, bold=True, color="C62828")
    cell.fill = RESULT_FILL
    cell.border = THIN_BORDER
    per_hit = f"B{R}"
    R += 1

    # Total damage (per hit × repeats)
    ws.cell(row=R, column=1, value="總傷害(含連擊)").font = LABEL_FONT
    ws.cell(row=R, column=1).border = THIN_BORDER
    cell = ws.cell(row=R, column=2, value=f"={per_hit}*MAX(1,{s_repeats})")
    cell.font = Font(name="Consolas", size=12, bold=True, color="C62828")
    cell.fill = RESULT_FILL
    cell.border = THIN_BORDER
    total_dmg = f"B{R}"
    R += 1

    # Damage / HP ratio
    ws.cell(row=R, column=1, value="傷害/敵HP比").font = LABEL_FONT
    ws.cell(row=R, column=1).border = THIN_BORDER
    cell = ws.cell(row=R, column=2, value=f"=IF({e_mhp}>0,{total_dmg}/{e_mhp},0)")
    cell.font = Font(name="Consolas", size=12, bold=True, color="1565C0")
    cell.number_format = '0.0%'
    cell.fill = RESULT_FILL
    cell.border = THIN_BORDER
    R += 1

    # Variance range
    ws.cell(row=R, column=1, value="變異範圍").font = LABEL_FONT
    ws.cell(row=R, column=1).border = THIN_BORDER
    cell = ws.cell(row=R, column=2,
                   value=f'=ROUND({per_hit}*(1-{s_variance}/100),0)&" ~ "&ROUND({per_hit}*(1+{s_variance}/100),0)')
    cell.font = NUM_FONT
    cell.fill = RESULT_FILL
    cell.border = THIN_BORDER
    R += 1

    # Hits to kill
    ws.cell(row=R, column=1, value="擊殺所需次數").font = LABEL_FONT
    ws.cell(row=R, column=1).border = THIN_BORDER
    cell = ws.cell(row=R, column=2,
                   value=f'=IF({total_dmg}>0,ROUNDUP({e_mhp}/{total_dmg},0),"∞")')
    cell.font = Font(name="Consolas", size=12, bold=True, color="388E3C")
    cell.fill = RESULT_FILL
    cell.border = THIN_BORDER
    R += 2

    # ── Explanation ────────────────────────────────────────────────
    ws.cell(row=R, column=1, value="【公式說明】").font = LABEL_FONT
    R += 1
    notes = [
        "標準傷害 = MAX(1, (外功×外功係數 - 敵外防×外防係數 + 內功×內功係數 - 敵內防×內防係數)",
        "        × (1 + (輕功-敵輕功) / 速度除數) + MAX(0, (福緣-敵福緣)×運勢係數))",
        "百分比傷害 = 敵氣血 × 百分比值",
        "HP比例傷害 = MAX(1, (HP/MHP × HP倍率 - 敵外防×外防係數) × 速度加成)",
        "最終傷害 = 基礎傷害 × 元素倍率 × 連擊數",
        "元素倍率: >100% = 弱點(紅), <100% = 抗性(綠), =100% = 普通",
        "※ 此計算假設 HP 全滿，不含 Boost/Break Shield 加成",
    ]
    for note in notes:
        ws.cell(row=R, column=1, value=note).font = Font(
            name="Microsoft JhengHei", size=9, color="666666")
        R += 1

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    for i in range(3, 11):
        ws.column_dimensions[get_column_letter(i)].width = 10
    ws.sheet_properties.tabColor = "FF6B6B"


# ═══════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════
def main():
    print("=== 萬法同歸 — 傷害計算表 v2 ===")
    print()

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("建立隱藏查詢表...")
    n = write_actor_params_lookup(wb)
    print(f"  _角色參數: {n} rows")
    n = write_equip_lookup(wb)
    print(f"  _裝備參數: {n} rows")
    n = write_enemy_lookup(wb)
    print(f"  _敵人參數: {n} rows")
    n = write_skill_coeff_lookup(wb)
    print(f"  _技能係數: {n} rows")

    print("\n建立資料表...")
    n = write_player_skills(wb)
    print(f"  [1] 我方技能: {n} rows")
    n = write_enemy_skills(wb)
    print(f"  [2] 敵方技能: {n} rows")
    n = write_lightfoot(wb)
    print(f"  [3] 輕功總覽: {n} rows")
    n = write_actor_stats(wb)
    print(f"  [4] 角色數值: {n} rows")
    n = write_enemy_stats(wb)
    print(f"  [5] 敵人數值: {n} rows")
    n = write_equipment(wb)
    print(f"  [6] 裝備總覽: {n} rows")
    n = write_element_rates(wb)
    print(f"  [7] 元素抗性: {n} rows")

    print("\n建立傷害計算機...")
    write_calculator(wb)
    print("  [8] 傷害計算機: done")

    # Reorder sheets: visible first, hidden last
    visible = ["我方技能", "敵方技能", "輕功總覽", "角色數值", "敵人數值",
               "裝備總覽", "元素抗性", "傷害計算機"]
    hidden = ["_角色參數", "_裝備參數", "_敵人參數", "_技能係數"]
    order = visible + hidden
    sheet_map = {ws.title: i for i, ws in enumerate(wb.worksheets)}
    wb._sheets = [wb.worksheets[sheet_map[name]] for name in order if name in sheet_map]

    print(f"\n儲存至 {OUT_PATH}...")
    wb.save(str(OUT_PATH))
    print("完成！")


if __name__ == "__main__":
    main()
