"""
Sync 序章 & 第一章 MD scripts → XLSX 劇本對話 sheet.
MD files are the source of truth. Deletes stale inserts from old script,
updates all dialogue rows, inserts new rows, and updates 世界觀 sheet.
"""
import openpyxl
import re
import sys
sys.stdout.reconfigure(encoding='utf-8')

# ─────────────────────────────────────────────────
# 1. Parse MD dialogue tables
# ─────────────────────────────────────────────────
def parse_md_rows(filepath):
    """Extract dialogue rows from MD file.
    Handles \\| (RPG Maker wait-for-input) inside cells."""
    results = []
    with open(filepath, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line.startswith('|') or not line.endswith('|'):
                continue
            # Replace \| with placeholder to avoid split issues
            tmp = line[1:-1].replace('\\|', '\x00')
            parts = [p.strip().replace('\x00', '\\|') for p in tmp.split('|')]
            # Only include dialogue rows (S0-001, S1-165a, etc.)
            if parts and re.match(r'S\d+-\d+[a-z]*$', parts[0]):
                results.append((parts[0], parts))
    return results

s0 = parse_md_rows('consilience-writer/序章_黃裳典籍.md')
s1 = parse_md_rows('consilience-writer/第一章_劃月風雲.md')
md_dict = {}
md_order = []
for sid, cols in s0 + s1:
    md_dict[sid] = cols
    md_order.append(sid)

print(f"MD: {len(s0)} prologue + {len(s1)} ch1 = {len(md_dict)} total rows")

# ─────────────────────────────────────────────────
# 2. Load XLSX
# ─────────────────────────────────────────────────
wb = openpyxl.load_workbook('docs/萬法同歸_設計文件.xlsx')
ws = wb['劇本對話']

header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
NCOLS = len(header)
xlsx = []
for row in ws.iter_rows(min_row=2, values_only=True):
    xlsx.append(list(row))

print(f"XLSX: {len(xlsx)} rows x {NCOLS} cols")

def seq(row):
    return str(row[0]).strip() if row[0] else None

def convert_val(v, col_idx):
    """Convert MD value for XLSX storage."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.replace('<br>', '\n')
        # V[1] column: convert pure numbers to int
        if col_idx == 3:
            try:
                return int(v)
            except ValueError:
                pass
    return v

# ─────────────────────────────────────────────────
# 3. Phase 1: Delete stale rows, update matched rows
# ─────────────────────────────────────────────────
result = []
deleted = []
updated = []

# Known old-script inserts to delete (not in current MD)
OLD_INSERTS = {'S0-003a', 'S0-007a', 'S0-084a', 'S1-073a', 'S1-208a'}

for row in xlsx:
    s = seq(row)
    if s and s in OLD_INSERTS:
        # Delete stale old-script inserts
        deleted.append(s)
    elif s and s in md_dict:
        # Update row from MD source of truth
        mc = md_dict[s]
        for i in range(min(len(mc), NCOLS)):
            row[i] = convert_val(mc[i], i)
        result.append(row)
        updated.append(s)
    else:
        # Keep all other rows untouched (side quests, other data)
        result.append(row)

print(f"\nPhase 1:")
print(f"  Deleted {len(deleted)} stale rows: {deleted}")
print(f"  Updated {len(updated)} rows")

# ─────────────────────────────────────────────────
# 4. Phase 2: Insert new MD rows not in XLSX
# ─────────────────────────────────────────────────
def rebuild_idx():
    return {seq(r): i for i, r in enumerate(result) if seq(r)}

idx = rebuild_idx()
inserted = []

for sid in md_order:
    if sid not in idx:
        mc = md_dict[sid]
        nr = [None] * NCOLS
        for i in range(min(len(mc), NCOLS)):
            nr[i] = convert_val(mc[i], i)

        # Copy 場景地點 from base row if col 7 is empty
        base = re.sub(r'[a-z]+$', '', sid)
        if base in idx and NCOLS > 7:
            base_row = result[idx[base]]
            if base_row[7] and not nr[7]:
                nr[7] = base_row[7]

        # Find insert position (after base row + any existing sub-rows)
        if base in idx:
            pos = idx[base] + 1
            while pos < len(result):
                ns = seq(result[pos])
                if ns and ns.startswith(base) and ns > base and ns < sid:
                    pos += 1
                else:
                    break
            result.insert(pos, nr)
            idx = rebuild_idx()
            inserted.append(sid)
        else:
            print(f"  WARNING: base {base} not found for {sid}")

print(f"\nPhase 2:")
print(f"  Inserted {len(inserted)} new rows: {inserted}")

# ─────────────────────────────────────────────────
# 5. Phase 3: Write back to 劇本對話
# ─────────────────────────────────────────────────
print(f"\nPhase 3: Writing {len(result)} rows to 劇本對話")

for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row_cells:
        cell.value = None

for i, rd in enumerate(result):
    for j, v in enumerate(rd):
        if v is not None:
            ws.cell(row=i + 2, column=j + 1, value=v)

# ─────────────────────────────────────────────────
# 6. Phase 4: Update 世界觀 sheet
# ─────────────────────────────────────────────────
print("\nPhase 4: Updating 世界觀")
ws2 = wb['世界觀']

new_entries = {
    '封印體系': (
        '上古先人以地脈之力鎮壓混沌之物的禁制。壁畫中被鑿去的符號是封印標記。\n'
        '封印正在碎裂——大地深處有某種沉眠之物開始呼吸。'
    ),
    '混沌之源': (
        '先於天地秩序的原初之力。既非氣，亦非瑪那，是不屬於任何一方的異物。\n'
        '不隨死亡消散，存在於被改造魔物的深層。比典籍更古老，比帝國更深的存在。'
    ),
    '瑪那': (
        '帝國魔法的能量來源。與東方的「氣」本質相異。\n'
        '帝國以理式將瑪那系統化為可操控的能量。\n'
        '魔物體內可同時偵測到瑪那、氣、以及更深處的「異物」三層。'
    ),
}

# Also update descriptions for existing entries that were exposition-heavy
updates = {
    '三教九流': (
        '東方文明的思想根基。三教（儒、釋、道）為精神支柱，'
        '九流為知識體系的百花齊放。\n'
        '三教無牆，九流無界——一人可兼修百家，不以為奇。\n'
        '劇本呈現：劍客與僧人共飲月色，道人同儒者齊染風霜。'
    ),
    '求真 vs 求善': (
        '帝國哲學偏向「求真」（理式解構萬物），東方思想偏向「求善」'
        '（武學承載人命與善念）。\n'
        '核心張力：求真不是壞事，但容不下善的真，不算真正的真。\n'
        '劇本呈現：墨汐若「想了解世界不是壞事吧？」→ 談笑「求真不是壞事，'
        '但他們的真，容不下我們的善。」'
    ),
}

existing = set()
for r in ws2.iter_rows(min_row=2, values_only=True):
    if r[0]:
        existing.add(str(r[0]).strip())

# Update existing
for term, desc in updates.items():
    if term in existing:
        for rc in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
            if rc[0].value and str(rc[0].value).strip() == term:
                rc[1].value = desc
                print(f"  Updated: {term}")
                break

# Add new
for term, desc in new_entries.items():
    if term in existing:
        for rc in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
            if rc[0].value and str(rc[0].value).strip() == term:
                rc[1].value = desc
                print(f"  Updated: {term}")
                break
    else:
        lr = ws2.max_row + 1
        ws2.cell(row=lr, column=1, value=term)
        ws2.cell(row=lr, column=2, value=desc)
        print(f"  Added: {term}")

# ─────────────────────────────────────────────────
# 7. Save
# ─────────────────────────────────────────────────
output = 'docs/萬法同歸_設計文件.xlsx'
wb.save(output)
print(f"\n{'='*40}")
print(f"DONE — saved to {output}")
print(f"劇本對話: {len(result)} rows")
print(f"  Deleted: {len(deleted)}")
print(f"  Updated: {len(updated)}")
print(f"  Inserted: {len(inserted)}")
