"""
萬法同歸 — 傷害計算 Excel 生成器
Generates a comprehensive damage calculator spreadsheet.
Supports multi-class (3 job variants per character).
"""
import json, re, math, sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

DATA_DIR = Path("C:/Consilience/Consilience/data")
OUT_PATH = Path("C:/Consilience/萬法同歸_傷害計算表.xlsx")

# ── Load all game data ──────────────────────────────────────────────
def load_json(name):
    with open(DATA_DIR / name, "r", encoding="utf-8") as f:
        return json.load(f)

skills_data  = load_json("Skills.json")
enemies_data = load_json("Enemies.json")
actors_data  = load_json("Actors.json")
classes_data = load_json("Classes.json")
weapons_data = load_json("Weapons.json")
armors_data  = load_json("Armors.json")
system_data  = load_json("System.json")

# ── Actor → multi-class mapping ─────────────────────────────────────
MAPPING_PATH = DATA_DIR / "actor_class_mapping.json"
if MAPPING_PATH.exists():
    with open(MAPPING_PATH, "r", encoding="utf-8") as f:
        _raw_mapping = json.load(f)
    # Convert string keys to int
    ACTOR_CLASS_MAP = {int(k): v for k, v in _raw_mapping.items()}
else:
    ACTOR_CLASS_MAP = {}

JOB_LABELS = {0: "均衡", 1: "攻擊", 2: "防禦"}

# ── Element names ───────────────────────────────────────────────────
raw_elements = system_data["elements"]
ELEMENTS = []
for e in raw_elements:
    c = re.sub(r'\\I\[\d+\]', '', e).strip()
    ELEMENTS.append(c if c else "(無)")

# Core elements are 1-16
CORE_ELEM_IDS = list(range(1, 17))
CORE_ELEM_NAMES = [ELEMENTS[i] for i in CORE_ELEM_IDS]

# ── Helper: parse traits for element rates ──────────────────────────
def get_element_rates(traits):
    """Return dict of {elementId: rate} from traits with code=22."""
    rates = {}
    for t in traits:
        if t["code"] == 22:
            eid = t["dataId"]
            # Multiple traits of same element multiply
            if eid in rates:
                rates[eid] *= t["value"]
            else:
                rates[eid] = t["value"]
    return rates

# ── Build actor stat tables ─────────────────────────────────────────
PARAM_NAMES = ["氣血","內力","外功","外防","內功","內防","輕功","福緣"]
CALC_LEVELS = [20, 50, 99]

class Battler:
    """Represents an actor or enemy with stats for damage calculation."""
    def __init__(self, name, bid, stats, elem_rates, is_enemy=False,
                 job_name="", job_index=-1):
        self.name = name
        self.id = bid
        self.hp = stats[0]   # Max HP
        self.mp = stats[1]   # Max MP
        self.atk = stats[2]
        self.df = stats[3]   # DEF
        self.mat = stats[4]
        self.mdf = stats[5]
        self.agi = stats[6]
        self.luk = stats[7]
        self.mhp = stats[0]
        self.mmp = stats[1]
        self.elem_rates = elem_rates  # {elemId: multiplier}
        self.is_enemy = is_enemy
        self.stats = stats
        self.job_name = job_name      # 職業名稱 (e.g. "劃月少主")
        self.job_index = job_index    # 0=A(均衡), 1=B(攻擊), 2=C(防禦)

    def elem_rate(self, elem_id):
        """Get damage multiplier for a given element."""
        if elem_id <= 0:
            return 1.0  # No element
        return self.elem_rates.get(elem_id, 1.0)

    @property
    def job_label(self):
        """Short label like '均衡' / '攻擊' / '防禦'."""
        return JOB_LABELS.get(self.job_index, "")

    @property
    def display_name(self):
        """Name with job variant for column headers."""
        if self.job_name:
            return f"{self.name}({self.job_name})"
        return self.name

def build_actors(level):
    """Build actor battlers at given level (variant A only, for backward compat)."""
    actors = []
    for a in actors_data:
        if not a:
            continue
        cid = a["classId"]
        cls = classes_data[cid]
        if not cls:
            continue
        idx = min(level - 1, len(cls["params"][0]) - 1)
        stats = [cls["params"][p][idx] for p in range(8)]

        # Merge class traits + actor traits for element rates
        all_traits = cls.get("traits", []) + a.get("traits", [])
        elem_rates = get_element_rates(all_traits)

        job_name = cls.get("name", "")
        actors.append(Battler(a["name"], a["id"], stats, elem_rates,
                              job_name=job_name, job_index=0))
    return actors

def build_actors_multiclass(level):
    """Build actor battlers at given level with all 3 job variants.

    Returns list of Battler objects sorted by (actor_id, job_index).
    Each actor appears 3 times (A/B/C variants).
    """
    actors = []
    for a in actors_data:
        if not a:
            continue
        aid = a["id"]
        mapping = ACTOR_CLASS_MAP.get(aid)
        if not mapping:
            # Fallback: no mapping, use default class
            cid = a["classId"]
            cls = classes_data[cid]
            if not cls:
                continue
            idx = min(level - 1, len(cls["params"][0]) - 1)
            stats = [cls["params"][p][idx] for p in range(8)]
            all_traits = cls.get("traits", []) + a.get("traits", [])
            elem_rates = get_element_rates(all_traits)
            actors.append(Battler(a["name"], aid, stats, elem_rates,
                                  job_name=cls.get("name", ""), job_index=0))
            continue

        class_ids = mapping["classes"]     # [classA_id, classB_id, classC_id]
        job_names = mapping["job_names"]   # [nameA, nameB, nameC]

        for ji, (cid, jname) in enumerate(zip(class_ids, job_names)):
            if cid >= len(classes_data) or not classes_data[cid]:
                continue
            cls = classes_data[cid]
            idx = min(level - 1, len(cls["params"][0]) - 1)
            stats = [cls["params"][p][idx] for p in range(8)]

            # Element rates: merge class traits + actor traits
            all_traits = cls.get("traits", []) + a.get("traits", [])
            elem_rates = get_element_rates(all_traits)

            actors.append(Battler(a["name"], aid, stats, elem_rates,
                                  job_name=jname, job_index=ji))
    return actors

def build_enemies():
    """Build enemy battlers."""
    enemies = []
    for e in enemies_data:
        if not e or not e.get("name") or e["name"].startswith("---"):
            continue
        stats = e["params"]
        elem_rates = get_element_rates(e.get("traits", []))
        enemies.append(Battler(e["name"], e["id"], stats, elem_rates, is_enemy=True))
    return enemies

# ── Damage formula evaluator ────────────────────────────────────────
class FormulaAttacker:
    """Proxy for 'a' in damage formulas."""
    def __init__(self, b):
        self.atk = b.atk
        self.mat = b.mat
        self.agi = b.agi
        self.luk = b.luk
        self.hp = b.hp
        self.mhp = b.mhp
        self.mp = b.mp
        self.mmp = b.mmp
        # Some formulas use a.def
        self.__dict__["def"] = b.df
        self.mdf = b.mdf

class FormulaDefender:
    """Proxy for 'b' in damage formulas."""
    def __init__(self, b):
        self.atk = b.atk
        self.mat = b.mat
        self.agi = b.agi
        self.luk = b.luk
        self.hp = b.hp
        self.mhp = b.mhp
        self.mp = b.mp
        self.mmp = b.mmp
        self.__dict__["def"] = b.df
        self.mdf = b.mdf

def eval_formula(formula_str, attacker, defender):
    """Evaluate a JS damage formula using Python."""
    if not formula_str or formula_str == "0":
        return 0

    a = FormulaAttacker(attacker)
    b = FormulaDefender(defender)

    # Transform JS to Python
    f = formula_str
    f = f.replace("Math.max", "max")
    f = f.replace("Math.min", "min")
    f = f.replace("Math.abs", "abs")
    f = f.replace("Math.floor", "int")
    f = f.replace("Math.ceil", "math.ceil")
    f = f.replace("Math.round", "round")
    f = f.replace("Math.random()", "0.5")
    f = f.replace("Math.PI", "3.14159")

    # Handle a.def and b.def (Python reserved word)
    f = re.sub(r'\ba\.def\b', 'a.__dict__["def"]', f)
    f = re.sub(r'\bb\.def\b', 'b.__dict__["def"]', f)

    # Handle 'power' variable (only skills 6,7 use it) - default to 1
    f = f.replace("power", "1")

    # Handle ternary operator: cond ? x : y → (x if cond else y)
    # Simple cases only
    f = re.sub(r'(\S+)\s*\?\s*(\S+)\s*:\s*(\S+)', r'(\2 if \1 else \3)', f)

    # Handle v[n] variables (game variables) - default to 0
    f = re.sub(r'v\[(\d+)\]', '0', f)
    f = re.sub(r'\$gameVariables\.value\(\d+\)', '0', f)

    # Handle user.xxx references
    f = f.replace("user.", "a.")
    f = f.replace("target.", "b.")

    try:
        result = eval(f, {"__builtins__": {}, "max": max, "min": min, "abs": abs,
                          "int": int, "round": round, "math": math,
                          "a": a, "b": b})
        return max(0, round(result)) if isinstance(result, (int, float)) else 0
    except Exception:
        return -1  # Formula evaluation failed

# ── Build skill data ────────────────────────────────────────────────
class SkillInfo:
    def __init__(self, sid, name, formula, elem_id, dtype, variance):
        self.id = sid
        self.name = name
        self.formula = formula
        self.elem_id = elem_id
        self.dtype = dtype  # 1=phys, 2=mag, 3=certain, 5=drain, 6=custom
        self.variance = variance
        self.elem_name = ELEMENTS[elem_id] if 0 < elem_id < len(ELEMENTS) else "(無)"

def build_skills():
    """Build list of all damage-dealing skills."""
    result = []
    for s in skills_data:
        if not s:
            continue
        dmg = s.get("damage", {})
        dtype = dmg.get("type", 0)
        formula = dmg.get("formula", "")
        if dtype <= 0 or not formula or formula == "0":
            continue
        name = s["name"]
        if name.startswith("---") or name.startswith("__") or name.startswith("－"):
            continue
        elem_id = dmg.get("elementId", 0)
        if elem_id < 0:
            elem_id = 0  # Treat normal-attack element as no element
        variance = dmg.get("variance", 20)
        result.append(SkillInfo(s["id"], name, formula, elem_id, dtype, variance))
    return result

# ── Map skills to owners ────────────────────────────────────────────
def map_enemy_skills():
    """Return {enemy_id: [skill_ids]} from enemy actions."""
    result = {}
    for e in enemies_data:
        if not e or not e.get("name") or e["name"].startswith("---"):
            continue
        sids = [act["skillId"] for act in e.get("actions", [])]
        if sids:
            result[e["id"]] = sids
    return result

def map_actor_skills():
    """Return {actor_id: [skill_ids]} from class learnings and skill separators."""
    result = {}

    # From class learnings — collect from ALL class variants
    for a in actors_data:
        if not a:
            continue
        aid = a["id"]
        sids = set()

        # Primary class
        cid = a["classId"]
        cls = classes_data[cid]
        if cls:
            for l in cls.get("learnings", []):
                sids.add(l["skillId"])

        # Additional class variants from mapping
        mapping = ACTOR_CLASS_MAP.get(aid)
        if mapping:
            for cid in mapping["classes"]:
                if cid < len(classes_data) and classes_data[cid]:
                    for l in classes_data[cid].get("learnings", []):
                        sids.add(l["skillId"])

        result[aid] = list(sids)

    # Name mapping: separator labels → actor IDs
    actor_name_map = {}
    for a in actors_data:
        if a:
            actor_name_map[a["name"]] = a["id"]

    # Manual alias mapping for non-obvious names
    ALIASES = {
        "東方啓": "東方啟",     # 啓→啟
        "煙菲花": "湮菲花",     # 煙→湮
        "染幽": "殷染幽",       # partial name
        "盼盼": "湮菲花",       # nickname
        "繾染": "殷染幽",       # alternate name
        "沫檸": "白沫檸",       # partial name
    }

    def resolve_owner(label):
        """Resolve a separator label to an actor ID."""
        label = label.strip()
        if label in actor_name_map:
            return actor_name_map[label]
        if label in ALIASES:
            resolved = ALIASES[label]
            if resolved in actor_name_map:
                return actor_name_map[resolved]
        for aname, aid in actor_name_map.items():
            if label in aname or aname in label:
                return aid
        return None

    # Parse BOTH separator styles
    current_owner = None
    empty_streak = 0
    for s in skills_data:
        if not s:
            continue
        name = s["name"]

        if name.startswith("---"):
            inner = name.strip("-").strip()
            if inner:
                resolved = resolve_owner(inner)
                if resolved is not None:
                    current_owner = resolved
                else:
                    current_owner = None
            empty_streak = 0
            continue

        if name.startswith("【") and name.endswith("】"):
            inner = name.strip("【】").strip()
            current_owner = resolve_owner(inner)
            empty_streak = 0
            continue

        dmg = s.get("damage", {})
        has_damage = dmg.get("type", 0) > 0 and dmg.get("formula", "") not in ("", "0")
        if not has_damage:
            empty_streak += 1
            if empty_streak >= 8:
                current_owner = None
            continue
        else:
            empty_streak = 0

        if current_owner is not None:
            if current_owner not in result:
                result[current_owner] = []
            if s["id"] not in result[current_owner]:
                result[current_owner].append(s["id"])

    return result

# ── Compute damage with element rate ────────────────────────────────
def calc_damage(skill, attacker, defender):
    """Calculate damage including element rate."""
    base = eval_formula(skill.formula, attacker, defender)
    if base <= 0:
        return base
    elem_rate = defender.elem_rate(skill.elem_id)
    return max(0, round(base * elem_rate))

# ── Excel styling ───────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="2B3A4A", end_color="2B3A4A", fill_type="solid")
HEADER_FONT = Font(name="Noto Sans TC", size=10, bold=True, color="FFFFFF")
SUB_HEADER_FILL = PatternFill(start_color="3D5A6E", end_color="3D5A6E", fill_type="solid")
DATA_FONT = Font(name="Noto Sans TC", size=9)
NUM_FONT = Font(name="Consolas", size=9)
ELEM_FILLS = {
    "劍法": "FFE0E0", "刀法": "FFD0B0", "棍法": "C8B090",
    "槍法": "D0D0FF", "拳掌": "FFE8C0", "音律": "E0D0F0",
    "奇門": "D0F0D0", "弓術": "F0E0C0", "筆法": "D8D8D8",
    "暗器": "B0B0B0", "短兵": "E0C0C0", "醫術": "C0F0C0",
    "毒術": "B0E0B0", "陰": "C0C0E0", "陽": "FFF0C0",
    "混元": "E0E0E0",
}
# Alternate row fills for job variants
JOB_FILLS = {
    0: None,  # A (均衡) — no fill
    1: PatternFill(start_color="FFF5E6", end_color="FFF5E6", fill_type="solid"),  # B (攻擊) warm
    2: PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid"),  # C (防禦) cool
}
THIN_BORDER = Border(
    left=Side(style="thin", color="C0C0C0"),
    right=Side(style="thin", color="C0C0C0"),
    top=Side(style="thin", color="C0C0C0"),
    bottom=Side(style="thin", color="C0C0C0"),
)

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

def style_data_cell(cell, is_num=False):
    cell.font = NUM_FONT if is_num else DATA_FONT
    cell.border = THIN_BORDER
    if is_num:
        cell.alignment = Alignment(horizontal="right")

# ── Sheet 1: 招式資料 ──────────────────────────────────────────────
def write_skills_sheet(wb, all_skills):
    ws = wb.active
    ws.title = "招式資料"

    headers = ["ID", "招式名稱", "傷害類型", "元素", "公式", "變異%"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    dtype_map = {1: "物理", 2: "魔法", 3: "必中", 4: "物理", 5: "吸收", 6: "自訂"}

    for i, sk in enumerate(all_skills, 2):
        ws.cell(row=i, column=1, value=sk.id)
        ws.cell(row=i, column=2, value=sk.name)
        ws.cell(row=i, column=3, value=dtype_map.get(sk.dtype, str(sk.dtype)))
        ws.cell(row=i, column=4, value=sk.elem_name)
        ws.cell(row=i, column=5, value=sk.formula)
        ws.cell(row=i, column=6, value=sk.variance)
        for col in range(1, 7):
            style_data_cell(ws.cell(row=i, column=col), col in (1, 6))

    ws.auto_filter.ref = f"A1:F{len(all_skills)+1}"
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 80
    ws.column_dimensions["F"].width = 8
    ws.freeze_panes = "A2"

# ── Sheet 2: 數值表（角色）— 3 rows per actor ─────────────────────
def write_actor_stats_sheet(wb, actors_mc_by_level):
    """Write actor stats with 3 rows per character (one per job variant)."""
    ws = wb.create_sheet("角色數值")

    headers = ["ID", "角色", "職業", "類型", "等級"] + PARAM_NAMES
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    row = 2
    for lv, actors in sorted(actors_mc_by_level.items()):
        for a in actors:
            ws.cell(row=row, column=1, value=a.id)
            ws.cell(row=row, column=2, value=a.name)
            ws.cell(row=row, column=3, value=a.job_name)
            ws.cell(row=row, column=4, value=a.job_label)
            ws.cell(row=row, column=5, value=lv)
            for p in range(8):
                ws.cell(row=row, column=6+p, value=a.stats[p])

            # Style all cells, apply job variant fill
            job_fill = JOB_FILLS.get(a.job_index)
            for col in range(1, 14):
                cell = ws.cell(row=row, column=col)
                style_data_cell(cell, col >= 5)
                if job_fill and col <= 4:
                    cell.fill = job_fill
            row += 1

    ws.auto_filter.ref = f"A1:M{row-1}"
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 6
    ws.column_dimensions["E"].width = 6
    for i in range(6, 14):
        ws.column_dimensions[get_column_letter(i)].width = 8
    ws.freeze_panes = "F2"

# ── Sheet 3: 數值表（敵人） ─────────────────────────────────────────
def write_enemy_stats_sheet(wb, enemies):
    ws = wb.create_sheet("敵人數值")

    headers = ["ID", "敵人"] + PARAM_NAMES
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    for i, e in enumerate(enemies, 2):
        ws.cell(row=i, column=1, value=e.id)
        ws.cell(row=i, column=2, value=e.name)
        for p in range(8):
            ws.cell(row=i, column=3+p, value=e.stats[p])
        for col in range(1, 11):
            style_data_cell(ws.cell(row=i, column=col), col >= 3)

    ws.auto_filter.ref = f"A1:J{len(enemies)+1}"
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 14
    for i in range(3, 11):
        ws.column_dimensions[get_column_letter(i)].width = 8
    ws.freeze_panes = "C2"

# ── Sheet 4: 元素抗性 ──────────────────────────────────────────────
def write_element_sheet(wb, actors_lv50, enemies):
    """Element rates — uses variant A only (rates are same across variants)."""
    ws = wb.create_sheet("元素抗性")

    headers = ["類型", "ID", "名稱"] + CORE_ELEM_NAMES
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    # Deduplicate actors (only show variant A)
    seen_ids = set()
    unique_actors = []
    for a in actors_lv50:
        if a.id not in seen_ids:
            seen_ids.add(a.id)
            unique_actors.append(a)

    row = 2
    for a in unique_actors:
        ws.cell(row=row, column=1, value="角色")
        ws.cell(row=row, column=2, value=a.id)
        ws.cell(row=row, column=3, value=a.name)
        for ci, eid in enumerate(CORE_ELEM_IDS):
            rate = a.elem_rate(eid)
            cell = ws.cell(row=row, column=4+ci, value=rate)
            cell.number_format = '0%'
            style_data_cell(cell, True)
            if rate < 0.5:
                cell.fill = PatternFill(start_color="C0FFC0", fill_type="solid")
            elif rate > 1.0:
                cell.fill = PatternFill(start_color="FFC0C0", fill_type="solid")
        for col in range(1, 4):
            style_data_cell(ws.cell(row=row, column=col))
        row += 1

    # Separator
    ws.cell(row=row, column=1, value="────────")
    row += 1

    # Enemies
    for e in enemies:
        ws.cell(row=row, column=1, value="敵人")
        ws.cell(row=row, column=2, value=e.id)
        ws.cell(row=row, column=3, value=e.name)
        for ci, eid in enumerate(CORE_ELEM_IDS):
            rate = e.elem_rate(eid)
            cell = ws.cell(row=row, column=4+ci, value=rate)
            cell.number_format = '0%'
            style_data_cell(cell, True)
            if rate < 0.5:
                cell.fill = PatternFill(start_color="C0FFC0", fill_type="solid")
            elif rate > 1.0:
                cell.fill = PatternFill(start_color="FFC0C0", fill_type="solid")
        for col in range(1, 4):
            style_data_cell(ws.cell(row=row, column=col))
        row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row-1}"
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 14
    for i in range(4, 4 + len(CORE_ELEM_IDS)):
        ws.column_dimensions[get_column_letter(i)].width = 8
    ws.freeze_panes = "D2"

# ── Sheet 5: 角色 vs 敵人 傷害表（含職業欄） ────────────────────────
def write_actor_vs_enemy_sheet(wb, actors_mc_lv50, enemies, all_skills, actor_skill_map):
    """Each row = (角色, 職業, 招式) — all 3 variants shown."""
    ws = wb.create_sheet("角色攻擊敵人")

    skill_by_id = {s.id: s for s in all_skills}

    fixed_cols = ["角色", "職業", "類型", "招式ID", "招式名稱", "元素", "傷害類型"]
    headers = fixed_cols + [f"{e.name}({e.id})" for e in enemies]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    # Sub-header row with enemy HP
    fc = len(fixed_cols)
    for ci in range(1, fc + 1):
        cell = ws.cell(row=2, column=ci, value="")
        cell.fill = SUB_HEADER_FILL
        cell.font = Font(name="Noto Sans TC", size=8, color="FFFFFF")
    ws.cell(row=2, column=fc, value="HP")
    for ci, e in enumerate(enemies):
        cell = ws.cell(row=2, column=fc+1+ci, value=e.hp)
        cell.fill = SUB_HEADER_FILL
        cell.font = Font(name="Consolas", size=8, color="FFFFFF")
        cell.alignment = Alignment(horizontal="right")

    dtype_map = {1: "物理", 2: "魔法", 3: "必中", 4: "物理", 5: "吸收", 6: "自訂"}

    row = 3
    total = 0
    for actor in actors_mc_lv50:
        sids = actor_skill_map.get(actor.id, [])
        if not sids:
            continue
        for sid in sids:
            sk = skill_by_id.get(sid)
            if not sk:
                continue
            ws.cell(row=row, column=1, value=actor.name)
            ws.cell(row=row, column=2, value=actor.job_name)
            ws.cell(row=row, column=3, value=actor.job_label)
            ws.cell(row=row, column=4, value=sk.id)
            ws.cell(row=row, column=5, value=sk.name)
            ws.cell(row=row, column=6, value=sk.elem_name)
            ws.cell(row=row, column=7, value=dtype_map.get(sk.dtype, "?"))

            for ci, enemy in enumerate(enemies):
                dmg = calc_damage(sk, actor, enemy)
                cell = ws.cell(row=row, column=fc+1+ci, value=dmg)
                style_data_cell(cell, True)
                if dmg > 0 and enemy.hp > 0:
                    ratio = dmg / enemy.hp
                    if ratio >= 1.0:
                        cell.fill = PatternFill(start_color="FF8080", fill_type="solid")
                    elif ratio >= 0.5:
                        cell.fill = PatternFill(start_color="FFCC80", fill_type="solid")
                    elif ratio >= 0.25:
                        cell.fill = PatternFill(start_color="FFFFB0", fill_type="solid")
                elif dmg < 0:
                    cell.fill = PatternFill(start_color="D0D0D0", fill_type="solid")
                    cell.value = "ERR"

            # Style fixed columns with job variant fill
            job_fill = JOB_FILLS.get(actor.job_index)
            for col in range(1, fc + 1):
                cell = ws.cell(row=row, column=col)
                style_data_cell(cell, col == 4)
                if job_fill and col <= 3:
                    cell.fill = job_fill

            row += 1
            total += 1

    print(f"  角色攻擊敵人: {total} rows x {len(enemies)} enemies")

    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{row-1}"
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 7
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 7
    ws.column_dimensions["G"].width = 7
    for i in range(fc + 1, fc + 1 + len(enemies)):
        ws.column_dimensions[get_column_letter(i)].width = 9
    ws.freeze_panes = f"{get_column_letter(fc+1)}3"

# ── Sheet 6: 敵人 vs 角色 傷害表（列頭含職業） ─────────────────────
def write_enemy_vs_actor_sheet(wb, actors_mc_lv50, enemies, all_skills, enemy_skill_map):
    """Column headers = 角色名(職業名) for all 3 variants."""
    ws = wb.create_sheet("敵人攻擊角色")

    skill_by_id = {s.id: s for s in all_skills}

    fixed_cols = ["敵人", "招式ID", "招式名稱", "元素", "類型"]
    # Column headers: one per actor+job variant
    headers = fixed_cols + [a.display_name for a in actors_mc_lv50]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    # Sub-header with actor HP and job variant color
    ws.cell(row=2, column=5, value="HP")
    for ci, a in enumerate(actors_mc_lv50):
        cell = ws.cell(row=2, column=6+ci, value=a.hp)
        cell.fill = SUB_HEADER_FILL
        cell.font = Font(name="Consolas", size=8, color="FFFFFF")
    for col in range(1, 6):
        cell = ws.cell(row=2, column=col)
        cell.fill = SUB_HEADER_FILL
        cell.font = Font(name="Noto Sans TC", size=8, color="FFFFFF")

    dtype_map = {1: "物理", 2: "魔法", 3: "必中", 4: "物理", 5: "吸收", 6: "自訂"}

    row = 3
    total = 0
    enemy_lookup = {e.id: e for e in enemies}

    for eid, sids in sorted(enemy_skill_map.items()):
        enemy = enemy_lookup.get(eid)
        if not enemy:
            continue
        for sid in sids:
            sk = skill_by_id.get(sid)
            if not sk:
                continue
            ws.cell(row=row, column=1, value=enemy.name)
            ws.cell(row=row, column=2, value=sk.id)
            ws.cell(row=row, column=3, value=sk.name)
            ws.cell(row=row, column=4, value=sk.elem_name)
            ws.cell(row=row, column=5, value=dtype_map.get(sk.dtype, "?"))

            for ci, actor in enumerate(actors_mc_lv50):
                dmg = calc_damage(sk, enemy, actor)
                cell = ws.cell(row=row, column=6+ci, value=dmg)
                style_data_cell(cell, True)
                if dmg > 0 and actor.hp > 0:
                    ratio = dmg / actor.hp
                    if ratio >= 1.0:
                        cell.fill = PatternFill(start_color="FF8080", fill_type="solid")
                    elif ratio >= 0.5:
                        cell.fill = PatternFill(start_color="FFCC80", fill_type="solid")
                    elif ratio >= 0.25:
                        cell.fill = PatternFill(start_color="FFFFB0", fill_type="solid")
                elif dmg < 0:
                    cell.fill = PatternFill(start_color="D0D0D0", fill_type="solid")
                    cell.value = "ERR"

            for col in range(1, 6):
                style_data_cell(ws.cell(row=row, column=col))

            row += 1
            total += 1

    print(f"  敵人攻擊角色: {total} rows x {len(actors_mc_lv50)} actor variants")

    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{row-1}"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 7
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 7
    ws.column_dimensions["E"].width = 6
    for i in range(6, 6 + len(actors_mc_lv50)):
        ws.column_dimensions[get_column_letter(i)].width = 12
    ws.freeze_panes = "F3"

# ── Sheet 7: 全技能 vs 全目標（快速查詢用） ─────────────────────────
def write_all_skills_vs_all_sheet(wb, actors_lv50, enemies, all_skills):
    """Write ALL skills evaluated with a generic attacker (average Lv50 stats)."""
    ws = wb.create_sheet("招式傷害速查")

    # Use average Lv50 stats as generic attacker (variant A only)
    seen_ids = set()
    unique_actors = []
    for a in actors_lv50:
        if a.id not in seen_ids and a.job_index == 0:
            seen_ids.add(a.id)
            unique_actors.append(a)

    avg_stats = [0] * 8
    for a in unique_actors:
        for i in range(8):
            avg_stats[i] += a.stats[i]
    if unique_actors:
        avg_stats = [s // len(unique_actors) for s in avg_stats]
    generic = Battler("平均Lv50", 0, avg_stats, {})

    targets = enemies

    fixed_cols = ["招式ID", "招式名稱", "元素", "類型"]
    headers = fixed_cols + [f"{t.name}" for t in targets]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    # HP row
    for ci, t in enumerate(targets):
        cell = ws.cell(row=2, column=5+ci, value=t.hp)
        cell.fill = SUB_HEADER_FILL
        cell.font = Font(name="Consolas", size=8, color="FFFFFF")
    for col in range(1, 5):
        cell = ws.cell(row=2, column=col)
        cell.fill = SUB_HEADER_FILL

    dtype_map = {1: "物理", 2: "魔法", 3: "必中", 4: "物理", 5: "吸收", 6: "自訂"}

    row = 3
    for sk in all_skills:
        ws.cell(row=row, column=1, value=sk.id)
        ws.cell(row=row, column=2, value=sk.name)
        ws.cell(row=row, column=3, value=sk.elem_name)
        ws.cell(row=row, column=4, value=dtype_map.get(sk.dtype, "?"))

        for ci, t in enumerate(targets):
            dmg = calc_damage(sk, generic, t)
            cell = ws.cell(row=row, column=5+ci, value=dmg)
            style_data_cell(cell, True)
            if dmg < 0:
                cell.value = "ERR"

        for col in range(1, 5):
            style_data_cell(ws.cell(row=row, column=col))
        row += 1

    print(f"  招式傷害速查: {len(all_skills)} skills x {len(targets)} targets")

    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{row-1}"
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 7
    ws.column_dimensions["D"].width = 6
    for i in range(5, 5 + len(targets)):
        ws.column_dimensions[get_column_letter(i)].width = 9
    ws.freeze_panes = "E3"

# ── Main ────────────────────────────────────────────────────────────
def main():
    print("Loading data...")
    all_skills = build_skills()
    print(f"  {len(all_skills)} damage skills")

    # Build multi-class actors at each level
    actors_mc_by_level = {}
    for lv in CALC_LEVELS:
        actors_mc_by_level[lv] = build_actors_multiclass(lv)
    actors_mc_lv50 = actors_mc_by_level[50]

    n_actors = len(set(a.id for a in actors_mc_lv50))
    n_variants = len(actors_mc_lv50)
    print(f"  {n_actors} actors x 3 variants = {n_variants} at 3 levels")

    enemies = build_enemies()
    print(f"  {len(enemies)} enemies")

    actor_skill_map = map_actor_skills()
    enemy_skill_map = map_enemy_skills()

    a_total = sum(len(v) for v in actor_skill_map.values())
    e_total = sum(len(v) for v in enemy_skill_map.values())
    print(f"  Actor skill mappings: {a_total}")
    print(f"  Enemy skill mappings: {e_total}")

    print("\nBuilding Excel...")
    wb = Workbook()

    write_skills_sheet(wb, all_skills)
    print("  [1/7] 招式資料 done")

    write_actor_stats_sheet(wb, actors_mc_by_level)
    print("  [2/7] 角色數值 done")

    write_enemy_stats_sheet(wb, enemies)
    print("  [3/7] 敵人數值 done")

    write_element_sheet(wb, actors_mc_lv50, enemies)
    print("  [4/7] 元素抗性 done")

    write_actor_vs_enemy_sheet(wb, actors_mc_lv50, enemies, all_skills, actor_skill_map)
    print("  [5/7] 角色攻擊敵人 done")

    write_enemy_vs_actor_sheet(wb, actors_mc_lv50, enemies, all_skills, enemy_skill_map)
    print("  [6/7] 敵人攻擊角色 done")

    write_all_skills_vs_all_sheet(wb, actors_mc_lv50, enemies, all_skills)
    print("  [7/7] 招式傷害速查 done")

    print(f"\nSaving to {OUT_PATH}...")
    wb.save(str(OUT_PATH))
    print("Done!")

if __name__ == "__main__":
    main()
